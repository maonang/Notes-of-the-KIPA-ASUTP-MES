#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Извлечение подписей к рисункам из PDF файлов.

Модуль предназначен для автоматического извлечения подписей к рисункам из PDF-файлов,
где подписи представлены в виде пар "русский текст / английский текст".
Поддерживает два типа расположения подписей:
- На одной строке с разделителем " / "
- На разных строках (русская и английская версии)

Версия: 3.2
"""

import logging
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Generator, Any

import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ============================================================================
# Конфигурация
# ============================================================================

OUTPUT_DIR: str = "results_excel"
LOG_FORMAT: str = "%(asctime)s [%(levelname)s] %(message)s"
LOG_DATE_FORMAT: str = "%H:%M:%S"

# Порог для автоматического переключения режима (0-1 рисунок)
AUTO_SWITCH_THRESHOLD: int = 1

# Разделитель между русской и английской частями: " / "
SEPARATOR_PATTERN: str = r'\s+/\s+'

# ============================================================================
# Регулярные выражения для режимов извлечения (только для режимов 2A, 2B, 2C)
# ============================================================================

# Режим 2A: разные строки, формат с "Рисунок/Figure"
PATTERNS_MODE_2A: List[re.Pattern] = [
    re.compile(
        r'(?:Figure|Рисунок)\s+(\d+)\s*[–—-]\s*([^\n]+)',
        re.UNICODE | re.IGNORECASE,
    ),
]

# Режим 2B: разные строки, формат без "Рисунок" (простая нумерация)
PATTERNS_MODE_2B: List[re.Pattern] = [
    re.compile(r'^(\d+)\.\s+([^\n]+)', re.UNICODE | re.MULTILINE),
]

# Режим 2C: разные строки, смешанный
PATTERNS_MODE_2C: List[re.Pattern] = [
    re.compile(r'(?:Рисунок|Figure)\s+(\d+)[^\n]*', re.UNICODE | re.IGNORECASE),
    re.compile(r'^(\d+)\.\s+([^\n]+)', re.UNICODE | re.MULTILINE),
]

# ============================================================================
# Модели данных
# ============================================================================

@dataclass
class Figure:
    """
    Представляет одну запись о рисунке.

    Attributes:
        page: Номер страницы, на которой найден рисунок
        number: Номер рисунка
        name1: Русское название (или первая часть)
        name2: Английское название (или вторая часть)
    """
    page: int
    number: int
    name1: str = ""
    name2: str = ""


@dataclass
class ExtractionResult:
    """
    Результат обработки одного PDF-файла.

    Attributes:
        pdf_path: Путь к обработанному PDF-файлу
        layout_type: Тип расположения (1 - одна строка, 2 - разные строки)
        format_type: Тип формата (1 - с "Рисунок/Figure", 2 - нумерация, 3 - смешанный)
        start_page: Страница, с которой начато извлечение
        auto_switched: Был ли автоматически переключен режим
        original_layout: Исходный тип расположения (до переключения)
        original_format: Исходный тип формата (до переключения)
        figures: Список найденных рисунков
        errors: Список ошибок
        warnings: Список предупреждений
    """
    pdf_path: Path
    layout_type: int
    format_type: int
    start_page: int = 1
    auto_switched: bool = False
    original_layout: Optional[int] = None
    original_format: Optional[int] = None
    figures: List[Figure] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def mode_key(self) -> str:
        """Возвращает ключ режима в формате '1A', '1B' и т.д."""
        return f"{self.layout_type}{chr(64 + self.format_type)}"

    @property
    def success(self) -> bool:
        """Возвращает True, если обработка прошла без ошибок."""
        return not self.errors

    @property
    def count(self) -> int:
        """Возвращает количество найденных рисунков."""
        return len(self.figures)


# ============================================================================
# Настройка логирования
# ============================================================================

def setup_logging() -> logging.Logger:
    """Настраивает и возвращает логгер."""
    logging.basicConfig(
        format=LOG_FORMAT,
        datefmt=LOG_DATE_FORMAT,
        level=logging.INFO,
    )
    return logging.getLogger(__name__)


logger: logging.Logger = setup_logging()


# ============================================================================
# Утилиты для обработки текста
# ============================================================================

# Регулярное выражение для удаления невидимых Unicode-символов
_INVISIBLE_RE: re.Pattern = re.compile(
    r"[\u00ad\u00a0"
    r"\u2000-\u200f"
    r"\u2028\u2029"
    r"\u202a-\u202f"
    r"\u2060\u2061"
    r"\ufeff]",
    re.UNICODE,
)


def normalize(text: str) -> str:
    """
    Очищает строку от невидимых Unicode-символов и лишних пробелов.

    Args:
        text: Исходная строка

    Returns:
        Очищенная строка
    """
    text = _INVISIBLE_RE.sub("", text)
    text = unicodedata.normalize("NFKC", text)
    return " ".join(text.split())


def iter_page_lines_from(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> Generator[Tuple[int, str], None, None]:
    """
    Генерирует строки текста с указанной страницы PDF.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Номер страницы, с которой начинать (1-based)

    Yields:
        Кортеж (номер_страницы, строка_текста)
    """
    total_pages = len(pdf_reader.pages)

    for page_num in range(start_page, total_pages + 1):
        page = pdf_reader.pages[page_num - 1]  # конвертация в 0-based
        text: str = page.extract_text() or ""
        for raw_line in text.split("\n"):
            yield page_num, raw_line


def get_sample_text(pdf_reader: PyPDF2.PdfReader, num_pages: int = 2) -> str:
    """
    Возвращает образец текста из первых страниц для диагностики.

    Args:
        pdf_reader: Объект PdfReader
        num_pages: Количество страниц для анализа

    Returns:
        Строка с образцом текста
    """
    sample: List[str] = []
    for page_num, page in enumerate(pdf_reader.pages[:num_pages], start=1):
        text: str = page.extract_text() or ""
        lines: List[str] = text.split("\n")[:10]
        sample.append(f"--- Страница {page_num} ---")
        sample.extend(lines)
        sample.append("")
    return "\n".join(sample)


def count_cyrillic_ratio(text: str) -> float:
    """
    Возвращает долю кириллических символов в тексте.

    Args:
        text: Анализируемый текст

    Returns:
        Доля кириллицы от 0.0 до 1.0
    """
    if not text:
        return 0.0

    text_clean: str = re.sub(r'[\d\s\.\,\;\:\?\!\(\)\[\]\{\}]', '', text)
    if not text_clean:
        return 0.0

    cyrillic: int = len(re.findall(r'[а-яА-ЯёЁ]', text_clean))
    return cyrillic / len(text_clean)


def count_latin_ratio(text: str) -> float:
    """
    Возвращает долю латинских символов в тексте.

    Args:
        text: Анализируемый текст

    Returns:
        Доля латиницы от 0.0 до 1.0
    """
    if not text:
        return 0.0

    text_clean: str = re.sub(r'[\d\s\.\,\;\:\?\!\(\)\[\]\{\}]', '', text)
    if not text_clean:
        return 0.0

    latin: int = len(re.findall(r'[a-zA-Z]', text_clean))
    return latin / len(text_clean)


def split_by_separator(text: str) -> Tuple[str, str]:
    """
    Разделяет строку на русскую и английскую части по разделителю '/'.

    Логика:
    1. Если в строке ровно один символ '/' - считаем его разделителем
    2. Если несколько '/' - ищем " / " (пробел-слеш-пробел) как основной разделитель
    3. Если " / " не найден - возвращаем весь текст как русскую часть

    Args:
        text: Исходная строка

    Returns:
        Кортеж (русская_часть, английская_часть)
    """
    # Подсчитываем количество слешей
    slash_count: int = text.count('/')

    # Случай 1: ровно один слеш - он и есть разделитель
    if slash_count == 1:
        parts = text.split('/')
        russian = parts[0].strip()
        english = parts[1].strip() if len(parts) > 1 else ""
        return russian, english

    # Случай 2: несколько слешей - ищем разделитель с пробелами
    if slash_count > 1:
        separator_match: Optional[re.Match] = re.search(SEPARATOR_PATTERN, text)
        if separator_match:
            sep_start: int = separator_match.start()
            sep_end: int = separator_match.end()
            russian = text[:sep_start].strip()
            english = text[sep_end:].strip()
            return russian, english

    # Нет слешей или не нашли разделитель
    return text, ""


def remove_leading_number(text: str) -> str:
    """
    Удаляет номер рисунка из начала строки, если он есть.

    Удаляет паттерны вида:
    - "1. " (цифра с точкой и пробелом)
    - "1 - " (цифра с пробелом, тире, пробелом)
    - "1 " (цифра с пробелом)

    Args:
        text: Исходная строка

    Returns:
        Строка без номера в начале
    """
    if not text:
        return text

    # Паттерны для удаления номера в начале строки
    patterns = [
        r'^\d+\.\s*',        # "1. ", "12. "
        r'^\d+\s*[–—-]\s*',  # "1 - ", "1- ", "1 — "
        r'^\d+\s+',          # "1 ", "12 "
    ]

    for pattern in patterns:
        text = re.sub(pattern, '', text, count=1)

    return text.strip()


def post_process_figure(figure: Figure) -> Figure:
    """
    Пост-обработка извлеченной фигуры.

    Выполняет:
    1. Удаление номера рисунка из начала русской и английской частей
    2. Нормализацию пробелов

    Args:
        figure: Объект Figure для обработки

    Returns:
        Обработанный объект Figure
    """
    # Удаляем номер из начала строк
    figure.name1 = remove_leading_number(figure.name1)
    figure.name2 = remove_leading_number(figure.name2)

    # Нормализация
    figure.name1 = normalize(figure.name1)
    figure.name2 = normalize(figure.name2)

    return figure


# ============================================================================
# Функции извлечения для режима 1 (одна строка)
# ============================================================================

def extract_mode_1A(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> List[Figure]:
    """
    Режим 1A: одна строка, формат 'Рисунок X / Figure X'.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Страница, с которой начинать извлечение

    Returns:
        Список найденных рисунков
    """
    found_nums: Set[int] = set()
    figures: List[Figure] = []

    for page_num, raw_line in iter_page_lines_from(pdf_reader, start_page):
        line: str = normalize(raw_line)

        # Разделяем русскую и английскую части
        russian_part, english_part = split_by_separator(line)

        if not russian_part and not english_part:
            continue

        # Ищем номер рисунка
        fig_num: Optional[int] = None

        # Поиск номера в русской части
        russian_match = re.search(r'Рисунок\s+(\d+)', russian_part, re.IGNORECASE)
        if russian_match:
            fig_num = int(russian_match.group(1))

        # Если не нашли, ищем в английской части
        if fig_num is None:
            english_match = re.search(r'Figure\s+(\d+)', english_part, re.IGNORECASE)
            if english_match:
                fig_num = int(english_match.group(1))

        # Если нашли номер и он новый
        if fig_num is not None and fig_num not in found_nums:
            fig = Figure(
                page=page_num,
                number=fig_num,
                name1=russian_part,
                name2=english_part,
            )
            fig = post_process_figure(fig)
            figures.append(fig)
            found_nums.add(fig_num)

    figures.sort(key=lambda f: f.number)
    return figures


def extract_mode_1B(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> List[Figure]:
    """
    Режим 1B: одна строка, формат '1. Текст / Текст' (простая нумерация).

    Args:
        pdf_reader: Объект PdfReader
        start_page: Страница, с которой начинать извлечение

    Returns:
        Список найденных рисунков
    """
    found_nums: Set[int] = set()
    figures: List[Figure] = []

    for page_num, raw_line in iter_page_lines_from(pdf_reader, start_page):
        line: str = normalize(raw_line)

        # Ищем номер в начале строки
        num_match = re.match(r'^(\d+)\.', line)
        if not num_match:
            continue

        fig_num = int(num_match.group(1))

        if fig_num in found_nums:
            continue

        # Разделяем русскую и английскую части
        russian_part, english_part = split_by_separator(line)

        # Если разделения не произошло, вся строка - русская часть
        if not russian_part:
            russian_part = line
            english_part = ""

        fig = Figure(
            page=page_num,
            number=fig_num,
            name1=russian_part,
            name2=english_part,
        )
        fig = post_process_figure(fig)
        figures.append(fig)
        found_nums.add(fig_num)

    figures.sort(key=lambda f: f.number)
    return figures


def extract_mode_1C(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> List[Figure]:
    """
    Режим 1C: одна строка, смешанный формат.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Страница, с которой начинать извлечение

    Returns:
        Список найденных рисунков
    """
    found_nums: Set[int] = set()
    figures: List[Figure] = []

    for page_num, raw_line in iter_page_lines_from(pdf_reader, start_page):
        line: str = normalize(raw_line)

        # Разделяем русскую и английскую части
        russian_part, english_part = split_by_separator(line)

        # Пробуем найти номер в формате "Рисунок X" или "Figure X"
        fig_num: Optional[int] = None

        # Поиск в разделенных частях
        if russian_part:
            match = re.search(r'(?:Рисунок|Figure)\s+(\d+)', russian_part, re.IGNORECASE)
            if match:
                fig_num = int(match.group(1))

        if fig_num is None and english_part:
            match = re.search(r'(?:Рисунок|Figure)\s+(\d+)', english_part, re.IGNORECASE)
            if match:
                fig_num = int(match.group(1))

        # Если не нашли, пробуем формат с нумерацией в начале
        if fig_num is None:
            num_match = re.match(r'^(\d+)\.', line)
            if num_match:
                fig_num = int(num_match.group(1))

        if fig_num is not None and fig_num not in found_nums:
            fig = Figure(
                page=page_num,
                number=fig_num,
                name1=russian_part if russian_part else line,
                name2=english_part,
            )
            fig = post_process_figure(fig)
            figures.append(fig)
            found_nums.add(fig_num)

    figures.sort(key=lambda f: f.number)
    return figures


# ============================================================================
# Функции извлечения для режима 2 (разные строки)
# ============================================================================

def extract_mode_2A(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> List[Figure]:
    """
    Режим 2A: разные строки, формат с 'Рисунок/Figure'.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Страница, с которой начинать извлечение

    Returns:
        Список найденных рисунков
    """
    temp_figures: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {"page": None, "name1": "", "name2": ""}
    )

    for page_num, raw_line in iter_page_lines_from(pdf_reader, start_page):
        line: str = normalize(raw_line)

        for pattern in PATTERNS_MODE_2A:
            match: Optional[re.Match] = pattern.search(line)
            if not match:
                continue

            fig_num: int = int(match.group(1))
            fig_text: str = normalize(match.group(2))

            if temp_figures[fig_num]["page"] is None:
                temp_figures[fig_num]["page"] = page_num

            # Определяем язык по наличию кириллицы
            if re.search(r'[а-яА-ЯёЁ]', fig_text):
                temp_figures[fig_num]["name1"] = fig_text
            else:
                temp_figures[fig_num]["name2"] = fig_text
            break

    figures: List[Figure] = []
    for fig_num, data in temp_figures.items():
        if data["page"] is not None:
            fig = Figure(
                page=data["page"],
                number=fig_num,
                name1=data["name1"],
                name2=data["name2"],
            )
            fig = post_process_figure(fig)
            figures.append(fig)

    figures.sort(key=lambda f: f.number)
    return figures


def extract_mode_2B(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> List[Figure]:
    """
    Режим 2B: разные строки, формат без 'Рисунок' (простая нумерация).

    Args:
        pdf_reader: Объект PdfReader
        start_page: Страница, с которой начинать извлечение

    Returns:
        Список найденных рисунков
    """
    temp_figures: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {"page": None, "name1": "", "name2": "", "count": 0}
    )

    for page_num, raw_line in iter_page_lines_from(pdf_reader, start_page):
        line: str = normalize(raw_line)

        for pattern in PATTERNS_MODE_2B:
            match: Optional[re.Match] = pattern.match(line)
            if not match:
                continue

            fig_num: int = int(match.group(1))
            fig_text: str = normalize(match.group(2))

            if temp_figures[fig_num]["page"] is None:
                temp_figures[fig_num]["page"] = page_num

            temp_figures[fig_num]["count"] += 1

            if re.search(r'[а-яА-ЯёЁ]', fig_text):
                temp_figures[fig_num]["name1"] = fig_text
            else:
                temp_figures[fig_num]["name2"] = fig_text
            break

    figures: List[Figure] = []
    for fig_num, data in temp_figures.items():
        if data["page"] is not None and data["count"] >= 1:
            fig = Figure(
                page=data["page"],
                number=fig_num,
                name1=data["name1"],
                name2=data["name2"],
            )
            fig = post_process_figure(fig)
            figures.append(fig)

    figures.sort(key=lambda f: f.number)
    return figures


def extract_mode_2C(pdf_reader: PyPDF2.PdfReader, start_page: int = 1) -> List[Figure]:
    """
    Режим 2C: разные строки, смешанный формат.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Страница, с которой начинать извлечение

    Returns:
        Список найденных рисунков
    """
    temp_figures: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {"page": None, "name1": "", "name2": "", "count": 0}
    )

    for page_num, raw_line in iter_page_lines_from(pdf_reader, start_page):
        line: str = normalize(raw_line)
        found: bool = False

        # Пробуем формат с "Рисунок/Figure"
        for pattern in PATTERNS_MODE_2A:
            match: Optional[re.Match] = pattern.search(line)
            if match:
                try:
                    fig_num: int = int(match.group(1))
                    fig_text: str = normalize(
                        match.group(2) if len(match.groups()) > 1 else ""
                    )

                    if temp_figures[fig_num]["page"] is None:
                        temp_figures[fig_num]["page"] = page_num

                    temp_figures[fig_num]["count"] += 1

                    if re.search(r'[а-яА-ЯёЁ]', fig_text):
                        temp_figures[fig_num]["name1"] = fig_text
                    else:
                        temp_figures[fig_num]["name2"] = fig_text
                    found = True
                    break
                except (IndexError, ValueError):
                    continue

        if found:
            continue

        # Пробуем формат с нумерацией
        for pattern in PATTERNS_MODE_2B:
            match = pattern.match(line)
            if match:
                try:
                    fig_num = int(match.group(1))
                    fig_text = normalize(match.group(2))

                    if temp_figures[fig_num]["page"] is None:
                        temp_figures[fig_num]["page"] = page_num

                    temp_figures[fig_num]["count"] += 1

                    if re.search(r'[а-яА-ЯёЁ]', fig_text):
                        temp_figures[fig_num]["name1"] = fig_text
                    else:
                        temp_figures[fig_num]["name2"] = fig_text
                    break
                except (IndexError, ValueError):
                    continue

    figures: List[Figure] = []
    for fig_num, data in temp_figures.items():
        if data["page"] is not None and data["count"] >= 1:
            fig = Figure(
                page=data["page"],
                number=fig_num,
                name1=data["name1"],
                name2=data["name2"],
            )
            fig = post_process_figure(fig)
            figures.append(fig)

    figures.sort(key=lambda f: f.number)
    return figures


# ============================================================================
# Основная функция извлечения
# ============================================================================

def extract_figures(
    pdf_path: Path,
    layout_type: int,
    format_type: int,
    start_page: int = 1,
    auto_switch: bool = True,
) -> ExtractionResult:
    """
    Извлекает рисунки из PDF-файла с возможностью автоматического переключения режима.

    Args:
        pdf_path: Путь к PDF-файлу
        layout_type: Тип расположения (1 - одна строка, 2 - разные строки)
        format_type: Тип формата (1 - с "Рисунок", 2 - нумерация, 3 - смешанный)
        start_page: Страница, с которой начинать извлечение
        auto_switch: Включить автоматическое переключение режима

    Returns:
        Объект ExtractionResult с результатами обработки
    """
    result = ExtractionResult(
        pdf_path=pdf_path,
        layout_type=layout_type,
        format_type=format_type,
        start_page=start_page,
        original_layout=layout_type,
        original_format=format_type,
    )

    if not pdf_path.exists():
        result.errors.append(f"Файл не найден: {pdf_path}")
        logger.error(result.errors[-1])
        return result

    layout_names: Dict[int, str] = {1: "одна строка с /", 2: "разные строки"}
    format_names: Dict[int, str] = {
        1: "с 'Рисунок/Figure'",
        2: "без 'Рисунок' (нумерация)",
        3: "смешанный",
    }

    logger.info(
        "Обработка: %s (расположение: %s, формат: %s, стартовая страница: %d)",
        pdf_path.name,
        layout_names[layout_type],
        format_names[format_type],
        start_page,
    )

    try:
        with pdf_path.open("rb") as fh:
            pdf_reader: PyPDF2.PdfReader = PyPDF2.PdfReader(fh)
            total_pages = len(pdf_reader.pages)
            logger.info("Страниц в файле: %d", total_pages)

            # Извлечение в выбранном режиме
            if layout_type == 1:
                if format_type == 1:
                    result.figures = extract_mode_1A(pdf_reader, start_page)
                elif format_type == 2:
                    result.figures = extract_mode_1B(pdf_reader, start_page)
                else:
                    result.figures = extract_mode_1C(pdf_reader, start_page)
            else:
                if format_type == 1:
                    result.figures = extract_mode_2A(pdf_reader, start_page)
                elif format_type == 2:
                    result.figures = extract_mode_2B(pdf_reader, start_page)
                else:
                    result.figures = extract_mode_2C(pdf_reader, start_page)

            # Автоматическое переключение режима при недостаточном количестве результатов
            if auto_switch and result.count <= AUTO_SWITCH_THRESHOLD:
                logger.warning(
                    "Найдено только %d рисунков. Пытаюсь найти более подходящий режим...",
                    result.count,
                )

                best_layout: int = layout_type
                best_format: int = format_type
                best_figures: List[Figure] = result.figures
                best_count: int = result.count

                # Перебор всех режимов
                for test_layout in (1, 2):
                    for test_format in (1, 2, 3):
                        if test_layout == layout_type and test_format == format_type:
                            continue

                        logger.info("Пробуем режим %d%c...", test_layout, chr(64 + test_format))

                        if test_layout == 1:
                            if test_format == 1:
                                test_raw = extract_mode_1A(pdf_reader, start_page)
                            elif test_format == 2:
                                test_raw = extract_mode_1B(pdf_reader, start_page)
                            else:
                                test_raw = extract_mode_1C(pdf_reader, start_page)
                        else:
                            if test_format == 1:
                                test_raw = extract_mode_2A(pdf_reader, start_page)
                            elif test_format == 2:
                                test_raw = extract_mode_2B(pdf_reader, start_page)
                            else:
                                test_raw = extract_mode_2C(pdf_reader, start_page)

                        if len(test_raw) > best_count:
                            best_layout = test_layout
                            best_format = test_format
                            best_figures = test_raw
                            best_count = len(test_raw)
                            logger.info(
                                "Режим %d%c дал больше результатов: %d рисунков",
                                test_layout,
                                chr(64 + test_format),
                                best_count,
                            )

                # Применяем более удачный режим, если он найден
                if (best_layout != layout_type or best_format != format_type) and best_count > result.count:
                    result.auto_switched = True
                    result.layout_type = best_layout
                    result.format_type = best_format
                    result.figures = best_figures

                    result.warnings.append(
                        f"Автоматически переключен: расположение {layout_names[layout_type]} → {layout_names[best_layout]}, "
                        f"формат {format_names[format_type]} → {format_names[best_format]}. "
                        f"Найдено рисунков: {best_count} (было: {result.count})"
                    )
                    logger.warning(result.warnings[-1])
                elif result.count == 0:
                    sample: str = get_sample_text(pdf_reader)
                    result.warnings.append(
                        f"Не найдено ни одного рисунка ни в одном режиме.\n"
                        f"Образец текста из PDF (первые 10 строк каждой страницы):\n{sample}"
                    )
                    logger.warning(result.warnings[-1])

    except PyPDF2.errors.PdfReadError as exc:
        result.errors.append(f"Повреждённый PDF: {exc}")
        logger.error(result.errors[-1])
    except Exception as exc:
        result.errors.append(f"Неожиданная ошибка: {exc}")
        logger.exception("Неожиданная ошибка при обработке %s", pdf_path.name)

    logger.info(
        "Итоговый результат: режим %d%c, найдено рисунков: %d",
        result.layout_type,
        chr(64 + result.format_type),
        result.count,
    )
    return result


# ============================================================================
# Сохранение результатов в Excel
# ============================================================================

def save_to_excel(result: ExtractionResult, output_dir: str = OUTPUT_DIR) -> Optional[Path]:
    """
    Сохраняет результат обработки в Excel-файл.

    Args:
        result: Результат обработки PDF
        output_dir: Директория для сохранения

    Returns:
        Путь к сохраненному файлу или None при ошибке
    """
    if not result.figures and not result.warnings:
        logger.warning("Нет данных для сохранения: %s", result.pdf_path.name)
        return None

    out_dir: Path = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    stem: str = result.pdf_path.stem
    timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path: Path = out_dir / f"{stem}_{timestamp}.xlsx"

    # Стили для Excel
    header_font: Font = Font(bold=True, size=11)
    header_fill: PatternFill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid",
    )
    header_font_white: Font = Font(bold=True, size=11, color="FFFFFF")

    note_font: Font = Font(italic=True, size=10, color="888888")
    warning_font: Font = Font(color="FF6600")

    thin_border: Border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    center_alignment: Alignment = Alignment(horizontal='center', vertical='center')
    left_alignment: Alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    try:
        wb: Workbook = Workbook()
        ws = wb.active
        ws.title = "Рисунки"

        # Заголовки таблицы
        headers: List[str] = ["Страница", "Номер рисунка", "Название (Русский)", "Название (English)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Данные
        current_row: int = 2
        for fig in result.figures:
            ws.cell(row=current_row, column=1, value=fig.page)
            ws.cell(row=current_row, column=2, value=fig.number)
            ws.cell(row=current_row, column=3, value=fig.name1)
            ws.cell(row=current_row, column=4, value=fig.name2)

            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col in (3, 4):
                    cell.alignment = left_alignment

            current_row += 1

        if result.figures:
            current_row += 1

        # Примечания
        note_row: int = current_row

        if result.auto_switched:
            layout_names: Dict[int, str] = {1: "одна строка с /", 2: "разные строки"}
            format_names: Dict[int, str] = {
                1: "с 'Рисунок/Figure'",
                2: "без 'Рисунок' (нумерация)",
                3: "смешанный",
            }

            cell = ws.cell(row=note_row, column=1, value="ПРИМЕЧАНИЕ:")
            cell.font = note_font
            cell = ws.cell(
                row=note_row,
                column=2,
                value=f"Автоматически переключен: "
                      f"{layout_names[result.original_layout]}, {format_names[result.original_format]} → "
                      f"{layout_names[result.layout_type]}, {format_names[result.format_type]}",
            )
            cell.font = note_font
            note_row += 1

        if result.warnings:
            for warning in result.warnings[:2]:
                cell = ws.cell(row=note_row, column=1, value="ПРЕДУПРЕЖДЕНИЕ:")
                cell.font = warning_font
                cell = ws.cell(row=note_row, column=2, value=warning[:200])
                cell.font = warning_font
                note_row += 1

        # Итоговая строка
        summary_row: int = max(note_row, current_row)
        cell = ws.cell(row=summary_row, column=1, value="ИТОГО:")
        cell.font = header_font
        cell = ws.cell(row=summary_row, column=2, value=f"Найдено рисунков: {result.count}")
        cell.font = header_font

        # Настройка ширины колонок
        column_widths: List[int] = [12, 15, 60, 60]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Заморозка заголовка
        ws.freeze_panes = 'A2'

        wb.save(out_path)
        logger.info("Excel файл сохранён: %s", out_path)
        return out_path

    except Exception as exc:
        logger.error("Не удалось сохранить Excel: %s", exc)
        return None


# ============================================================================
# Поиск PDF-файлов
# ============================================================================

def find_pdf_files(directory: str = ".") -> List[Path]:
    """
    Находит все PDF-файлы в указанной директории.

    Args:
        directory: Путь к директории для поиска

    Returns:
        Отсортированный список путей к PDF-файлам
    """
    return sorted(Path(directory).glob("*.pdf"))


def ask_start_page(total_pages: int) -> int:
    """
    Запрашивает у пользователя номер страницы, с которой начинать извлечение.

    Args:
        total_pages: Общее количество страниц в PDF

    Returns:
        Номер стартовой страницы (1-based)
    """
    print("\n" + "=" * 70)
    print("С КАКОЙ СТРАНИЦЫ НАЧАТЬ ИЗВЛЕЧЕНИЕ?")
    print("=" * 70)
    print(f"  Всего страниц в файле: {total_pages}")
    print("=" * 70)

    while True:
        choice = input(f"\nВведите номер страницы (1-{total_pages}) или Enter для страницы 1: ").strip()

        if choice == "":
            return 1

        try:
            page_num = int(choice)
            if 1 <= page_num <= total_pages:
                return page_num
            else:
                print(f"Некорректный ввод. Введите число от 1 до {total_pages}")
        except ValueError:
            print("Некорректный ввод. Введите целое число.")


# ============================================================================
# Интерфейс командной строки
# ============================================================================

def ask_layout() -> int:
    """
    Запрашивает у пользователя тип расположения подписей.

    Returns:
        Выбранный тип (1 или 2)
    """
    print("\n" + "=" * 70)
    print("КАК РАСПОЛОЖЕНЫ ПОДПИСИ К РИСУНКАМ?")
    print("=" * 70)
    print("  1 — На одной строке с разделителем '/'")
    print("  2 — На разных строках (русский и английский)")
    print("=" * 70)

    while True:
        choice: str = input("\nВведите 1 или 2: ").strip()
        if choice in ("1", "2"):
            return int(choice)
        print("Некорректный ввод, попробуйте снова.")


def ask_format(layout_type: int) -> int:
    """
    Запрашивает у пользователя формат записи подписей.

    Args:
        layout_type: Тип расположения (влияет на примеры)

    Returns:
        Выбранный формат (1, 2 или 3)
    """
    print("\n" + "=" * 70)
    if layout_type == 1:
        print("КАКОЙ ФОРМАТ ЗАПИСИ НА СТРОКЕ?")
    else:
        print("КАКОЙ ФОРМАТ ЗАПИСИ НА СТРОКАХ?")
    print("=" * 70)
    print("  1 — Со словами 'Рисунок' и 'Figure'")
    print("  2 — Простая нумерация")
    print("  3 — Автоматическое определение")
    print("=" * 70)

    while True:
        choice: str = input("\nВведите 1, 2 или 3: ").strip()
        if choice in ("1", "2", "3"):
            return int(choice)
        print("Некорректный ввод, попробуйте снова.")


def main() -> None:
    """Главная функция программы."""
    # Выбор режимов
    layout_type: int = ask_layout()
    format_type: int = ask_format(layout_type)

    # Поиск PDF-файлов
    pdf_files: List[Path] = find_pdf_files()
    if not pdf_files:
        logger.warning("PDF файлы не найдены в текущей папке.")
        return

    # Для первого файла запрашиваем стартовую страницу
    first_pdf = pdf_files[0]
    with open(first_pdf, "rb") as fh:
        reader = PyPDF2.PdfReader(fh)
        total_pages = len(reader.pages)

    start_page: int = ask_start_page(total_pages)

    # Автоматическое переключение
    auto_switch_input: str = input(
        "\nВключить автоматическое определение формата при отсутствии результатов? (Y/n): "
    ).strip().lower()
    auto_switch: bool = auto_switch_input in ('', 'y', 'yes', 'д', 'да')

    logger.info("Найдено PDF файлов: %d", len(pdf_files))

    # Статистика
    total_figures: int = 0
    saved_files: List[Path] = []
    failed_files: List[Path] = []
    switched_files: List[Tuple[Path, str, str]] = []

    layout_names: Dict[int, str] = {1: "одна строка с /", 2: "разные строки"}
    format_names: Dict[int, str] = {
        1: "с 'Рисунок/Figure'",
        2: "без 'Рисунок' (нумерация)",
        3: "смешанный",
    }

    # Обработка каждого файла
    for idx, pdf_path in enumerate(pdf_files, start=1):
        logger.info("[%d/%d] %s", idx, len(pdf_files), pdf_path.name)

        result: ExtractionResult = extract_figures(
            pdf_path,
            layout_type,
            format_type,
            start_page,
            auto_switch,
        )

        if result.errors:
            failed_files.append(pdf_path)
            continue

        if result.auto_switched:
            switched_files.append((
                pdf_path,
                f"{layout_names[result.original_layout]}, {format_names[result.original_format]}",
                f"{layout_names[result.layout_type]}, {format_names[result.format_type]}",
            ))

        out_path: Optional[Path] = save_to_excel(result)
        if out_path:
            saved_files.append(out_path)
            total_figures += result.count

    # Итоговый отчет
    print("\n" + "=" * 70)
    print("  ОТЧЕТ ОБ ОБРАБОТКЕ")
    print("=" * 70)
    print(f"  Обработано файлов        : {len(pdf_files)}")
    print(f"  Успешно сохранено        : {len(saved_files)}")
    print(f"  Всего рисунков           : {total_figures}")

    if switched_files:
        print(f"\n  Автоматически переключено ({len(switched_files)} файлов):")
        for pdf_path, old_mode, new_mode in switched_files:
            print(f"    • {pdf_path.name}")
            print(f"      {old_mode} → {new_mode}")

    if failed_files:
        print(f"\n  Ошибки ({len(failed_files)}):")
        for p in failed_files:
            print(f"    • {p.name}")

    print("=" * 70)


if __name__ == "__main__":
    main()