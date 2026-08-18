#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Извлечение подписей к рисункам из PDF файлов.

Модуль предназначен для автоматического извлечения подписей к рисункам из PDF-файлов,
где подписи представлены в виде пар "русский текст / английский текст".
Поддерживает два типа расположения подписей:
- На одной строке с разделителем " / "
- На разных строках (русская и английская версии)
"""

import logging
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Generator, Any, Union, TypeAlias

import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ============================================================================
# Конфигурация
# ============================================================================

OUTPUT_DIR: str = "results_excel"
LOG_FORMAT: str = "%(asctime)s [%(levelname)s] %(message)s"
LOG_DATE_FORMAT: str = "%H:%M:%S"

AUTO_SWITCH_THRESHOLD: int = 1
SEPARATOR_PATTERN: str = r'\s+/\s+'
DEFAULT_SEPARATOR: str = " - "


# ============================================================================
# Перечисления
# ============================================================================

class LayoutType(Enum):
    """Тип расположения подписей."""
    ONE_LINE = 1
    MULTI_LINE = 2


class FormatType(Enum):
    """Тип формата записи подписей."""
    WITH_FIGURE = 1
    NUMBERING_ONLY = 2
    MIXED = 3


class Language(Enum):
    """Определение языка текста."""
    RUSSIAN = "russian"
    ENGLISH = "english"
    UNKNOWN = "unknown"


# ============================================================================
# Модели данных
# ============================================================================

@dataclass(frozen=True)
class Figure:
    """
    Представляет запись о рисунке.

    Attributes:
        page: Номер страницы
        number: Номер рисунка
        name_russian: Русское название
        name_english: Английское название
    """
    page: int
    number: int
    name_russian: str = ""
    name_english: str = ""


@dataclass
class ExtractionResult:
    """
    Результат обработки PDF-файла.

    Attributes:
        pdf_path: Путь к файлу
        layout_type: Тип расположения
        format_type: Тип формата
        start_page: Начальная страница
        separator: Разделитель
        auto_switched: Флаг автоматического переключения
        original_layout: Исходный тип расположения
        original_format: Исходный тип формата
        figures: Список найденных рисунков
        errors: Список ошибок
        warnings: Список предупреждений
    """
    pdf_path: Path
    layout_type: LayoutType
    format_type: FormatType
    start_page: int = 1
    separator: str = DEFAULT_SEPARATOR
    auto_switched: bool = False
    original_layout: Optional[LayoutType] = None
    original_format: Optional[FormatType] = None
    figures: List[Figure] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def mode_key(self) -> str:
        """Возвращает ключ режима в формате '1A', '1B' и т.д."""
        return f"{self.layout_type.value}{chr(64 + self.format_type.value)}"

    @property
    def success(self) -> bool:
        """Возвращает True, если обработка прошла без ошибок."""
        return not self.errors

    @property
    def count(self) -> int:
        """Возвращает количество найденных рисунков."""
        return len(self.figures)


# ============================================================================
# Настройка логирования
# ============================================================================

def _setup_logging() -> logging.Logger:
    """Настраивает и возвращает логгер."""
    logging.basicConfig(
        format=LOG_FORMAT,
        datefmt=LOG_DATE_FORMAT,
        level=logging.INFO,
    )
    return logging.getLogger(__name__)


logger: logging.Logger = _setup_logging()

# ============================================================================
# Утилиты для обработки текста
# ============================================================================

_INVISIBLE_RE: re.Pattern = re.compile(
    r"[\u00ad\u00a0"
    r"\u2000-\u200f"
    r"\u2028\u2029"
    r"\u202a-\u202f"
    r"\u2060\u2061"
    r"\ufeff]",
    re.UNICODE,
)


def _normalize_figure_markers(text: str) -> str:
    """
    Нормализует маркеры рисунков в тексте.

    Заменяет ошибочные написания на правильные
    """
    # Замена латинских букв на кириллические в слове "РИСУНОК"
    # P → Р, C → С, Y → У, H → Н, O → О, K → К
    replacements = {
        'PICYHOK': 'РИСУНОК',
        'PИСУНОК': 'РИСУНОК',  # P вместо Р
        'РИСУНОK': 'РИСУНОК',  # K вместо К
        'PИСУНОK': 'РИСУНОК',  # P вместо Р и K вместо К
    }

    for wrong, correct in replacements.items():
        text = text.replace(wrong, correct)

    return text


def normalize_text(text: str) -> str:
    """
    Очищает строку от невидимых Unicode-символов и нормализует пробелы.

    Args:
        text: Исходная строка

    Returns:
        Очищенная строка
    """
    if not text:
        return ""

    text = _INVISIBLE_RE.sub("", text)
    text = unicodedata.normalize("NFKC", text)
    text = _normalize_hyphens(text)
    text = _normalize_figure_markers(text)
    return " ".join(text.split())


def _normalize_hyphens(text: str) -> str:
    """
    Нормализует пробелы вокруг дефисов в кодах.

    Убирает пробелы вокруг дефисов в кодовых обозначениях:
    """
    if not text:
        return text

    for _ in range(3):
        text = re.sub(
            r'([A-Za-z0-9]+)\s*[–—-]\s*([A-Za-z0-9]+)',
            r'\1-\2',
            text
        )

    return text


def split_by_separator(text: str) -> Tuple[str, str]:
    """
    Разделяет строку на русскую и английскую части по разделителю '/'.

    Args:
        text: Исходная строка

    Returns:
        Кортеж (русская_часть, английская_часть)
    """
    if not text:
        return "", ""

    slash_count = text.count('/')

    if slash_count == 1:
        parts = text.split('/')
        return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""

    if slash_count > 1:
        match = re.search(SEPARATOR_PATTERN, text)
        if match:
            return text[:match.start()].strip(), text[match.end():].strip()

    return text, ""


def _iter_page_lines(
        pdf_reader: PyPDF2.PdfReader,
        start_page: int = 1
) -> Generator[Tuple[int, str], None, None]:
    """
    Генерирует строки текста со страниц PDF.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Номер страницы (1-based)

    Yields:
        Кортеж (номер_страницы, строка_текста)
    """
    for page_num in range(start_page, len(pdf_reader.pages) + 1):
        page = pdf_reader.pages[page_num - 1]
        text = page.extract_text() or ""
        for line in text.split("\n"):
            yield page_num, line


def get_sample_text(pdf_reader: PyPDF2.PdfReader, num_pages: int = 2) -> str:
    """
    Возвращает образец текста из первых страниц для диагностики.

    Args:
        pdf_reader: Объект PdfReader
        num_pages: Количество страниц для анализа

    Returns:
        Строка с образцом текста
    """
    sample = []
    for page_num in range(min(num_pages, len(pdf_reader.pages))):
        page = pdf_reader.pages[page_num]
        text = page.extract_text() or ""
        lines = text.split("\n")[:10]
        sample.append(f"--- Страница {page_num + 1} ---")
        sample.extend(lines)
        sample.append("")
    return "\n".join(sample)


# ============================================================================
# Паттерны для извлечения
# ============================================================================

def _compile_patterns() -> Dict[str, List[re.Pattern]]:
    """Компилирует регулярные выражения для извлечения подписей."""
    return {
        "with_figure": [
            re.compile(
                r'(?:Figure|Рисунок)\s*(\d+)\s*[–—-]\s*(.+)',
                re.UNICODE | re.IGNORECASE,
            ),
        ],
        "numbering_only": [
            re.compile(r'^(\d+)\.\s+([^\n]+)', re.UNICODE | re.MULTILINE),
        ],
        "mixed": [
            re.compile(
                r'(?:Рисунок|Figure)\s+(\d+)[^\n]*',
                re.UNICODE | re.IGNORECASE,
            ),
            re.compile(r'^(\d+)\.\s+([^\n]+)', re.UNICODE | re.MULTILINE),
        ],
    }


PATTERNS = _compile_patterns()


# ============================================================================
# Функции извлечения для режима "одна строка"
# ============================================================================

def _extract_one_line_with_figure(
        pdf_reader: PyPDF2.PdfReader,
        start_page: int
) -> List[Figure]:
    """
    Извлекает подписи в формате: "Рисунок X / Figure X".

    Args:
        pdf_reader: Объект PdfReader
        start_page: Начальная страница

    Returns:
        Список найденных рисунков
    """
    found_numbers = set()
    figures = []

    for page_num, raw_line in _iter_page_lines(pdf_reader, start_page):
        line = normalize_text(raw_line)

        russian_part, english_part = split_by_separator(line)
        if not russian_part and not english_part:
            continue

        figure_num = None

        russian_match = re.search(r'Рисунок\s+(\d+)', russian_part, re.IGNORECASE)
        if russian_match:
            figure_num = int(russian_match.group(1))

        if figure_num is None:
            english_match = re.search(r'Figure\s+(\d+)', english_part, re.IGNORECASE)
            if english_match:
                figure_num = int(english_match.group(1))

        if figure_num is not None and figure_num not in found_numbers:
            figures.append(Figure(
                page=page_num,
                number=figure_num,
                name_russian=russian_part,
                name_english=english_part,
            ))
            found_numbers.add(figure_num)

    return sorted(figures, key=lambda f: f.number)


def _extract_one_line_numbering_only(
        pdf_reader: PyPDF2.PdfReader,
        start_page: int
) -> List[Figure]:
    """
    Извлекает подписи в формате: "1. Текст / Текст".

    Args:
        pdf_reader: Объект PdfReader
        start_page: Начальная страница

    Returns:
        Список найденных рисунков
    """
    found_numbers = set()
    figures = []

    for page_num, raw_line in _iter_page_lines(pdf_reader, start_page):
        line = normalize_text(raw_line)

        match = re.match(r'^(\d+)\.', line)
        if not match:
            continue

        figure_num = int(match.group(1))
        if figure_num in found_numbers:
            continue

        russian_part, english_part = split_by_separator(line)
        if not russian_part:
            russian_part = line

        figures.append(Figure(
            page=page_num,
            number=figure_num,
            name_russian=russian_part,
            name_english=english_part,
        ))
        found_numbers.add(figure_num)

    return sorted(figures, key=lambda f: f.number)


def _extract_one_line_mixed(
        pdf_reader: PyPDF2.PdfReader,
        start_page: int
) -> List[Figure]:
    """
    Извлекает подписи в смешанном формате.

    Args:
        pdf_reader: Объект PdfReader
        start_page: Начальная страница

    Returns:
        Список найденных рисунков
    """
    found_numbers = set()
    figures = []

    for page_num, raw_line in _iter_page_lines(pdf_reader, start_page):
        line = normalize_text(raw_line)

        russian_part, english_part = split_by_separator(line)

        figure_num = None

        if russian_part:
            match = re.search(r'(?:Рисунок|Figure)\s+(\d+)', russian_part, re.IGNORECASE)
            if match:
                figure_num = int(match.group(1))

        if figure_num is None and english_part:
            match = re.search(r'(?:Рисунок|Figure)\s+(\d+)', english_part, re.IGNORECASE)
            if match:
                figure_num = int(match.group(1))

        if figure_num is None:
            match = re.match(r'^(\d+)\.', line)
            if match:
                figure_num = int(match.group(1))

        if figure_num is not None and figure_num not in found_numbers:
            figures.append(Figure(
                page=page_num,
                number=figure_num,
                name_russian=russian_part if russian_part else line,
                name_english=english_part,
            ))
            found_numbers.add(figure_num)

    return sorted(figures, key=lambda f: f.number)


# ============================================================================
# Функции извлечения для режима "разные строки"
# ============================================================================

def _extract_multi_line_with_figure(
        pdf_reader: PyPDF2.PdfReader,
        start_page: int
) -> List[Figure]:
    """
    Извлекает подписи на разных строках с маркерами "Рисунок/Figure".
    """
    temp_data: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {"page": None, "name_russian": "", "name_english": ""}
    )

    for page_num, raw_line in _iter_page_lines(pdf_reader, start_page):
        line = normalize_text(raw_line)

        for pattern in PATTERNS["with_figure"]:
            match = pattern.search(line)
            if not match:
                continue

            figure_num = int(match.group(1))
            figure_text = normalize_text(match.group(2))

            if temp_data[figure_num]["page"] is None:
                temp_data[figure_num]["page"] = page_num

            # Определяем язык ПО МАРКЕРУ в строке, а не по содержанию текста
            if 'РИСУНОК' in line.upper():
                temp_data[figure_num]["name_russian"] = figure_text
            elif 'FIGURE' in line.upper():
                temp_data[figure_num]["name_english"] = figure_text
            break

    figures = []
    for num, data in temp_data.items():
        if data["page"] is not None:
            figures.append(Figure(
                page=data["page"],
                number=num,
                name_russian=data["name_russian"],
                name_english=data["name_english"],
            ))

    return sorted(figures, key=lambda f: f.number)


def _extract_multi_line_numbering_only(
    pdf_reader: PyPDF2.PdfReader,
    start_page: int
) -> List[Figure]:
    """
    Извлекает подписи на разных строках с простой нумерацией.
    """
    temp_data: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {"page": None, "name_russian": "", "name_english": "", "count": 0}
    )

    for page_num, raw_line in _iter_page_lines(pdf_reader, start_page):
        line = normalize_text(raw_line)

        for pattern in PATTERNS["numbering_only"]:
            match = pattern.match(line)
            if not match:
                continue

            figure_num = int(match.group(1))
            figure_text = normalize_text(match.group(2))

            if temp_data[figure_num]["page"] is None:
                temp_data[figure_num]["page"] = page_num

            temp_data[figure_num]["count"] += 1

            # Простая проверка на наличие кириллицы
            if re.search(r'[а-яА-ЯёЁ]', figure_text):
                temp_data[figure_num]["name_russian"] = figure_text
            else:
                temp_data[figure_num]["name_english"] = figure_text
            break

    figures = []
    for num, data in temp_data.items():
        if data["page"] is not None and data["count"] >= 1:
            figures.append(Figure(
                page=data["page"],
                number=num,
                name_russian=data["name_russian"],
                name_english=data["name_english"],
            ))

    return sorted(figures, key=lambda f: f.number)


def _extract_multi_line_mixed(
    pdf_reader: PyPDF2.PdfReader,
    start_page: int
) -> List[Figure]:
    """
    Извлекает подписи на разных строках в смешанном формате.
    """
    temp_data: Dict[int, Dict[str, Any]] = defaultdict(
        lambda: {"page": None, "name_russian": "", "name_english": "", "count": 0}
    )

    for page_num, raw_line in _iter_page_lines(pdf_reader, start_page):
        line = normalize_text(raw_line)
        found = False

        for pattern in PATTERNS["with_figure"] + PATTERNS["mixed"]:
            match = pattern.search(line) if pattern in PATTERNS["with_figure"] else pattern.match(line)
            if not match:
                continue

            try:
                figure_num = int(match.group(1))
                figure_text = normalize_text(match.group(2) if len(match.groups()) > 1 else "")

                if temp_data[figure_num]["page"] is None:
                    temp_data[figure_num]["page"] = page_num

                temp_data[figure_num]["count"] += 1

                # Определяем язык ПО МАРКЕРУ в строке
                if 'РИСУНОК' in line.upper():
                    temp_data[figure_num]["name_russian"] = figure_text
                elif 'FIGURE' in line.upper():
                    temp_data[figure_num]["name_english"] = figure_text

                found = True
                break
            except (IndexError, ValueError):
                continue

        if found:
            continue

    figures = []
    for num, data in temp_data.items():
        if data["page"] is not None and data["count"] >= 1:
            figures.append(Figure(
                page=data["page"],
                number=num,
                name_russian=data["name_russian"],
                name_english=data["name_english"],
            ))

    return sorted(figures, key=lambda f: f.number)


# ============================================================================
# Извлечение данных
# ============================================================================

_EXTRACTION_FUNCTIONS = {
    (LayoutType.ONE_LINE, FormatType.WITH_FIGURE): _extract_one_line_with_figure,
    (LayoutType.ONE_LINE, FormatType.NUMBERING_ONLY): _extract_one_line_numbering_only,
    (LayoutType.ONE_LINE, FormatType.MIXED): _extract_one_line_mixed,
    (LayoutType.MULTI_LINE, FormatType.WITH_FIGURE): _extract_multi_line_with_figure,
    (LayoutType.MULTI_LINE, FormatType.NUMBERING_ONLY): _extract_multi_line_numbering_only,
    (LayoutType.MULTI_LINE, FormatType.MIXED): _extract_multi_line_mixed,
}


def _extract_figures(
        pdf_reader: PyPDF2.PdfReader,
        layout_type: LayoutType,
        format_type: FormatType,
        start_page: int
) -> List[Figure]:
    """
    Извлекает рисунки в соответствии с выбранным режимом.

    Args:
        pdf_reader: Объект PdfReader
        layout_type: Тип расположения
        format_type: Тип формата
        start_page: Начальная страница

    Returns:
        Список найденных рисунков
    """
    key = (layout_type, format_type)
    func = _EXTRACTION_FUNCTIONS.get(key)
    if func is None:
        return []
    return func(pdf_reader, start_page)


# ============================================================================
# Автоматическое переключение режима
# ============================================================================

def _try_switch_mode(
        pdf_reader: PyPDF2.PdfReader,
        current_layout: LayoutType,
        current_format: FormatType,
        start_page: int,
        current_figures: List[Figure]
) -> Tuple[LayoutType, FormatType, List[Figure]]:
    """
    Пытается найти более подходящий режим извлечения.

    Args:
        pdf_reader: Объект PdfReader
        current_layout: Текущий тип расположения
        current_format: Текущий тип формата
        start_page: Начальная страница
        current_figures: Текущий список рисунков

    Returns:
        Кортеж (новый_тип_расположения, новый_тип_формата, новый_список_рисунков)
    """
    best_layout = current_layout
    best_format = current_format
    best_figures = current_figures
    best_count = len(current_figures)

    for layout in LayoutType:
        for format_type in FormatType:
            if layout == current_layout and format_type == current_format:
                continue

            test_figures = _extract_figures(pdf_reader, layout, format_type, start_page)
            if len(test_figures) > best_count:
                best_layout = layout
                best_format = format_type
                best_figures = test_figures
                best_count = len(test_figures)

    return best_layout, best_format, best_figures


# ============================================================================
# Основная функция извлечения
# ============================================================================

def extract_figures(
        pdf_path: Path,
        layout_type: LayoutType,
        format_type: FormatType,
        start_page: int = 1,
        auto_switch: bool = True,
        separator: str = DEFAULT_SEPARATOR,
) -> ExtractionResult:
    """
    Извлекает рисунки из PDF-файла.

    Args:
        pdf_path: Путь к PDF-файлу
        layout_type: Тип расположения
        format_type: Тип формата
        start_page: Начальная страница
        auto_switch: Включить автоматическое переключение
        separator: Разделитель между номером и названием

    Returns:
        Результат обработки
    """
    result = ExtractionResult(
        pdf_path=pdf_path,
        layout_type=layout_type,
        format_type=format_type,
        start_page=start_page,
        separator=separator,
        original_layout=layout_type,
        original_format=format_type,
    )

    if not pdf_path.exists():
        result.errors.append(f"Файл не найден: {pdf_path}")
        logger.error(result.errors[-1])
        return result

    logger.info(
        "Обработка: %s (расположение: %s, формат: %s, страница: %d)",
        pdf_path.name,
        layout_type.name,
        format_type.name,
        start_page,
    )

    try:
        with pdf_path.open("rb") as fh:
            pdf_reader = PyPDF2.PdfReader(fh)
            logger.info("Всего страниц: %d", len(pdf_reader.pages))

            result.figures = _extract_figures(pdf_reader, layout_type, format_type, start_page)

            if auto_switch and result.count <= AUTO_SWITCH_THRESHOLD:
                logger.warning(
                    "Найдено только %d рисунков. Поиск более подходящего режима...",
                    result.count
                )

                new_layout, new_format, new_figures = _try_switch_mode(
                    pdf_reader, layout_type, format_type, start_page, result.figures
                )

                if (new_layout != layout_type or new_format != format_type) and len(new_figures) > result.count:
                    result.auto_switched = True
                    result.layout_type = new_layout
                    result.format_type = new_format
                    result.figures = new_figures

                    result.warnings.append(
                        f"Переключен режим: {layout_type.name}→{new_layout.name}, "
                        f"{format_type.name}→{new_format.name}. "
                        f"Найдено: {len(new_figures)} (было: {result.count})"
                    )
                    logger.warning(result.warnings[-1])
                elif result.count == 0:
                    sample = get_sample_text(pdf_reader)
                    result.warnings.append(
                        f"Рисунки не найдены ни в одном режиме.\n"
                        f"Образец текста:\n{sample}"
                    )
                    logger.warning(result.warnings[-1])

    except PyPDF2.errors.PdfReadError as exc:
        result.errors.append(f"Поврежденный PDF: {exc}")
        logger.error(result.errors[-1])
    except Exception as exc:
        result.errors.append(f"Ошибка: {exc}")
        logger.exception("Ошибка при обработке %s", pdf_path.name)

    logger.info(
        "Результат: режим %s%s, найдено: %d",
        result.layout_type.value,
        chr(64 + result.format_type.value),
        result.count
    )
    return result


# ============================================================================
# Сохранение в Excel
# ============================================================================

def _create_excel_styles():
    """Создает стили для Excel."""
    return {
        "header_font": Font(bold=True, size=11, color="FFFFFF"),
        "header_fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        "note_font": Font(italic=True, size=10, color="888888"),
        "warning_font": Font(color="FF6600"),
        "border": Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        ),
        "center": Alignment(horizontal='center', vertical='center'),
        "left": Alignment(horizontal='left', vertical='center', wrap_text=True),
    }


def save_to_excel(result: ExtractionResult, output_dir: str = OUTPUT_DIR) -> Optional[Path]:
    """
    Сохраняет результат обработки в Excel-файл.

    Args:
        result: Результат обработки
        output_dir: Директория для сохранения

    Returns:
        Путь к сохраненному файлу или None при ошибке
    """
    if not result.figures and not result.warnings:
        logger.warning("Нет данных для сохранения: %s", result.pdf_path.name)
        return None

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"{result.pdf_path.stem}_{timestamp}.xlsx"

    styles = _create_excel_styles()

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Рисунки"

        headers = ["Страница", "Номер рисунка", "Название (Русский)", "Название (English)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]

        row = 2
        for fig in result.figures:
            ws.cell(row=row, column=1, value=fig.page)
            ws.cell(row=row, column=2, value=fig.number)
            ws.cell(row=row, column=3, value=fig.name_russian)
            ws.cell(row=row, column=4, value=fig.name_english)

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = styles["border"]
                if col in (3, 4):
                    cell.alignment = styles["left"]

            row += 1

        if result.figures:
            row += 1

        note_row = row

        if result.auto_switched:
            cell = ws.cell(row=note_row, column=1, value="ПРИМЕЧАНИЕ:")
            cell.font = styles["note_font"]
            cell = ws.cell(
                row=note_row,
                column=2,
                value=f"Автоматически переключен: {result.original_layout.name}→{result.layout_type.name}, "
                      f"{result.original_format.name}→{result.format_type.name}"
            )
            cell.font = styles["note_font"]
            note_row += 1

        if result.warnings:
            for warning in result.warnings[:2]:
                cell = ws.cell(row=note_row, column=1, value="ПРЕДУПРЕЖДЕНИЕ:")
                cell.font = styles["warning_font"]
                cell = ws.cell(row=note_row, column=2, value=warning[:200])
                cell.font = styles["warning_font"]
                note_row += 1

        summary_row = max(note_row, row)
        cell = ws.cell(row=summary_row, column=1, value="ИТОГО:")
        cell.font = Font(bold=True, size=11)
        cell = ws.cell(row=summary_row, column=2, value=f"Найдено рисунков: {result.count}")
        cell.font = Font(bold=True, size=11)

        widths = [12, 15, 60, 60]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A2'
        wb.save(out_path)
        logger.info("Сохранено: %s", out_path)
        return out_path

    except Exception as exc:
        logger.error("Ошибка сохранения Excel: %s", exc)
        return None


# ============================================================================
# Интерфейс командной строки
# ============================================================================

def find_pdf_files(directory: str = ".") -> List[Path]:
    """Находит все PDF-файлы в директории."""
    return sorted(Path(directory).glob("*.pdf"))


def ask_start_page(total_pages: int) -> int:
    """Запрашивает у пользователя начальную страницу."""
    print("\n" + "=" * 70)
    print("С КАКОЙ СТРАНИЦЫ НАЧАТЬ ИЗВЛЕЧЕНИЕ?")
    print("=" * 70)

    while True:
        choice = input(f"\nВведите номер страницы (1-{total_pages}) или Enter для страницы 1: ").strip()
        if choice == "":
            return 1
        try:
            page_num = int(choice)
            if 1 <= page_num <= total_pages:
                return page_num
            print(f"Введите число от 1 до {total_pages}")
        except ValueError:
            print("Введите целое число.")


def ask_layout() -> LayoutType:
    """Запрашивает у пользователя тип расположения."""
    print("\n" + "=" * 70)
    print("КАК РАСПОЛОЖЕНЫ ПОДПИСИ К РИСУНКАМ?")
    print("=" * 70)
    print("  1 — На одной строке с разделителем '/'")
    print("  2 — На разных строках")
    print("=" * 70)

    while True:
        choice = input("\nВведите 1 или 2: ").strip()
        if choice == "1":
            return LayoutType.ONE_LINE
        elif choice == "2":
            return LayoutType.MULTI_LINE
        print("Некорректный ввод.")


def ask_format() -> FormatType:
    """Запрашивает у пользователя тип формата."""
    print("\n" + "=" * 70)
    print("КАКОЙ ФОРМАТ ЗАПИСИ?")
    print("=" * 70)
    print("  1 — Со словами 'Рисунок' и 'Figure'")
    print("  2 — Простая нумерация")
    print("  3 — Смешанный")
    print("=" * 70)

    while True:
        choice = input("\nВведите 1, 2 или 3: ").strip()
        if choice == "1":
            return FormatType.WITH_FIGURE
        elif choice == "2":
            return FormatType.NUMBERING_ONLY
        elif choice == "3":
            return FormatType.MIXED
        print("Некорректный ввод.")


def ask_separator() -> str:
    """Запрашивает у пользователя разделитель."""
    print("\n" + "=" * 70)
    print("РАЗДЕЛИТЕЛЬ МЕЖДУ НОМЕРОМ И НАЗВАНИЕМ?")
    print("=" * 70)
    print("  По умолчанию: ' - ' (пробел, тире, пробел)")
    print("=" * 70)

    choice = input("\nВведите разделитель (Enter для ' - '): ").strip()
    return choice if choice else " - "


def ask_auto_switch() -> bool:
    """Запрашивает у пользователя авто-переключение."""
    choice = input(
        "\nВключить автоматическое определение формата при отсутствии результатов? (Y/n): "
    ).strip().lower()
    return choice in ('н', '', 'y', 'yes', 'д', 'да')


def check_empty_fields(result: ExtractionResult) -> Dict[str, int]:
    """
    Проверяет наличие пустых полей в извлеченных рисунках.

    Args:
        result: Результат обработки

    Returns:
        Словарь со статистикой пустых полей
    """
    stats = {
        "empty_russian": 0,
        "empty_english": 0,
        "total": len(result.figures)
    }

    empty_russian_examples = []
    empty_english_examples = []

    for fig in result.figures:
        if not fig.name_russian or not fig.name_russian.strip():
            stats["empty_russian"] += 1
            if len(empty_russian_examples) < 3:
                empty_russian_examples.append(f"№{fig.number} (стр. {fig.page})")

        if not fig.name_english or not fig.name_english.strip():
            stats["empty_english"] += 1
            if len(empty_english_examples) < 3:
                empty_english_examples.append(f"№{fig.number} (стр. {fig.page})")

    # Добавляем примеры для отчета
    stats["russian_examples"] = empty_russian_examples
    stats["english_examples"] = empty_english_examples

    return stats


def print_empty_fields_report(stats: Dict[str, int]) -> None:
    """
    Выводит отчет о пустых полях в консоль.

    Args:
        stats: Статистика пустых полей
    """
    if stats["empty_russian"] == 0 and stats["empty_english"] == 0:
        print(f"Все поля отчета заполнены.")
        return

    print(f"\nОбнаружены пустые поля!")

    if stats["empty_russian"] > 0:
        print(f"    • RU: {stats['empty_russian']} из {stats['total']}")

    if stats["empty_english"] > 0:
        print(f"    • EN: {stats['empty_english']} из {stats['total']}")


def check_all_files_empty_fields(results: List[ExtractionResult]) -> Dict[str, int]:
    """
    Проверяет пустые поля для всех обработанных файлов.

    Args:
        results: Список результатов обработки

    Returns:
        Словарь с общей статистикой
    """
    total_stats = {
        "total_files": len(results),
        "total_figures": 0,
        "empty_russian_total": 0,
        "empty_english_total": 0,
        "files_with_empty": 0
    }

    for result in results:
        stats = check_empty_fields(result)
        total_stats["total_figures"] += stats["total"]
        total_stats["empty_russian_total"] += stats["empty_russian"]
        total_stats["empty_english_total"] += stats["empty_english"]
        if stats["empty_russian"] > 0 or stats["empty_english"] > 0:
            total_stats["files_with_empty"] += 1

    return total_stats


def main() -> None:
    layout_type = ask_layout()
    format_type = ask_format()
    separator = ask_separator()

    pdf_files = find_pdf_files()
    if not pdf_files:
        logger.warning("PDF файлы не найдены.")
        return

    with open(pdf_files[0], "rb") as fh:
        reader = PyPDF2.PdfReader(fh)
        total_pages = len(reader.pages)

    start_page = ask_start_page(total_pages)
    auto_switch = ask_auto_switch()

    logger.info("Найдено PDF файлов: %d", len(pdf_files))

    total_figures = 0
    saved_files = []
    failed_files = []
    switched_files = []
    all_results = []

    for idx, pdf_path in enumerate(pdf_files, start=1):
        logger.info("[%d/%d] %s", idx, len(pdf_files), pdf_path.name)

        result = extract_figures(
            pdf_path,
            layout_type,
            format_type,
            start_page,
            auto_switch,
            separator,
        )

        all_results.append(result)

        if result.errors:
            failed_files.append(pdf_path)
            continue

        if result.auto_switched:
            switched_files.append((
                pdf_path,
                f"{result.original_layout.name} / {result.original_format.name}",
                f"{result.layout_type.name} / {result.format_type.name}",
            ))

        out_path = save_to_excel(result)
        if out_path:
            saved_files.append(out_path)
            total_figures += result.count

    # Проверяем пустые поля для этого файла
    stats = check_empty_fields(result)
    print_empty_fields_report(stats)

    print("\n" + "=" * 70)
    print("  ОТЧЕТ ОБ ОБРАБОТКЕ")
    print("=" * 70)
    print(f"  Успешно обработано и сохранено : {len(saved_files)} / {len(pdf_files)}")
    print(f"  Всего рисунков                 : {total_figures}")

    if switched_files:
        print(f"\n  Автоматически переключено ({len(switched_files)} файлов):")
        for pdf_path, old_mode, new_mode in switched_files:
            print(f"    • {pdf_path.name}")
            print(f"      {old_mode} → {new_mode}")

    if failed_files:
        print(f"\n  Ошибки ({len(failed_files)}):")
        for p in failed_files:
            print(f"    • {p.name}")

    print("=" * 70)


if __name__ == "__main__":
    main()