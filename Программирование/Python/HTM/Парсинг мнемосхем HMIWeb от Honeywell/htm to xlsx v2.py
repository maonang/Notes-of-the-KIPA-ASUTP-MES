from __future__ import annotations
import argparse
import concurrent.futures
import json
import logging
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Any

from bs4 import BeautifulSoup

try:
    import lxml
    DEFAULT_PARSER = "lxml"
except Exception:
    DEFAULT_PARSER = "html.parser"

import openpyxl

# --- Настройка логирования ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)


# ---------- Структуры данных ----------
@dataclass
class HMIElement:
    """Одна запись элемента, найденного в HTM (до объединения с SHA)."""
    htm_file: Path
    element_id: str
    element_class: str
    src: str  # как указано в HTM (относительный/абсолютный)
    link_type: str
    hdxproperties_raw: str
    parameters_raw: str
    hdx_parsed: Dict[str, str] = field(default_factory=dict)
    parameters_parsed: Dict[str, str] = field(default_factory=dict)


@dataclass
class SHAMetadata:
    """Структура, которая хранит извлечённые метаданные из SHA."""
    abs_path: Path
    title: str = ""
    description: str = ""
    width: Optional[str] = None
    height: Optional[str] = None
    parameters_list: List[Dict[str, str]] = field(default_factory=list)
    scripts_list: List[Dict[str, Any]] = field(default_factory=list)

    def to_flat_dict(self) -> Dict[str, Any]:
        """Возвращает сплющенную репрезентацию для записи в таблицу."""
        return {
            "Файл SHAPE (SRC)": str(self.abs_path),
            "Title": self.title,
            "Description": self.description,
            "Width": (self.width or ""),
            "Height": (self.height or ""),
            "Параметры (Count)": len(self.parameters_list),
            "Параметры (Name, Type, Default)": json.dumps(self.parameters_list, ensure_ascii=False),
            "Скрипты (Count)": len(self.scripts_list),
        }


# ---------- Утилиты парсинга атрибутов ----------
def parse_hdxproperties(hdx_str: Optional[str]) -> Dict[str, str]:
    """Разбирает строку вида 'key:value;key2:value2;' в словарь."""
    result: Dict[str, str] = {}
    if not hdx_str:
        return result
    for item in filter(None, hdx_str.split(';')):
        parts = item.split(':', 1)
        if len(parts) == 2:
            key = parts[0].strip()
            value = parts[1].strip()
            if key:
                result[key] = value
    return result


def parse_parameters_attribute(parameters_str: Optional[str]) -> Dict[str, str]:
    """
    Разбирает атрибут parameters формата 'Type?Key:Value;...' и возвращает {Key: Value}.
    Обрабатывает нестандартные случаи аккуратно.
    """
    result: Dict[str, str] = {}
    if not parameters_str:
        return result
    # разделяем по ';'
    for item in filter(None, parameters_str.split(';')):
        # пробуем найти последний разделитель ':' который отделяет ключ от значения
        if ':' in item:
            left, right = item.rsplit(':', 1)
            # ключ может быть вида "Type?Key" — берем часть после '?'
            if '?' in left:
                _, key = left.split('?', 1)
            else:
                key = left
            key = key.strip()
            value = right.strip()
            if key:
                result[key] = value
        else:
            # Если нет ':', возможно вид "Key" или "Type?Key" - ставим пустое значение
            if '?' in item:
                _, key = item.split('?', 1)
            else:
                key = item
            key = key.strip()
            if key:
                result[key] = ""
    return result


# ---------- Парсер SHA ----------
class SHAParser:
    """
    Класс для разбора SHА-файлов (.sha) с возможностью параллельного запуска.
    Разбирает XML-подобный файл и собирает Title, Description, Width, Height,
    список параметров и список скриптов.
    """

    def __init__(self, parser: str = DEFAULT_PARSER):
        self.parser = parser

    @staticmethod
    def _read_with_fallback(path: Path) -> Optional[str]:
        """Пробуем читать SHA в нескольких кодировках."""
        encodings = ["windows-1252", "windows-1251", "utf-8", "latin-1"]
        for enc in encodings:
            try:
                return path.read_text(encoding=enc)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logging.debug("Ошибка чтения %s с кодировкой %s: %s", path, enc, e)
                break
        logging.error("Не удалось прочитать SHА-файл %s", path)
        return None

    def parse(self, sha_abs_path: Path) -> Optional[SHAMetadata]:
        """Разбирает один SHA и возвращает SHAMetadata или None при ошибке."""
        content = self._read_with_fallback(sha_abs_path)
        if content is None:
            return None

        try:
            soup = BeautifulSoup(content, "xml")
            shapefile = soup.find("shapefile")
            if not shapefile:
                # Иногда файлы могут быть не строго XML — пробуем искать по тегам менее строго
                shapefile = soup.find(lambda tag: tag.name and tag.name.lower() == "shapefile")
            if not shapefile:
                logging.warning("Формат SHА не содержит <shapefile>: %s", sha_abs_path)
                # возвращаем базовый объект с только путем
                return SHAMetadata(abs_path=sha_abs_path)

            # Извлекаем базовые атрибуты
            title = shapefile.get("title", "") or ""
            description = shapefile.get("description", "") or ""
            width = shapefile.get("width", "")
            height = shapefile.get("height", "")

            # Параметры
            parameters_list = []
            for param in shapefile.find_all("parameter"):
                parameters_list.append({
                    "Name": param.get("name", ""),
                    "Type": param.get("type", ""),
                    "Description": param.get("description", ""),
                    "DefaultValue": param.get("defaultvalue", "")
                })

            # Скрипты
            scripts_list = []
            for script in shapefile.find_all("script"):
                scripts_list.append({
                    "ID": script.get("id", "N/A"),
                    "Language": script.get("language", "N/A"),
                    "Event": script.get("event", "N/A"),
                    "HasCode": bool(script.text and script.text.strip()),
                    "CodeSnippet": (script.text or "").strip()
                })

            meta = SHAMetadata(
                abs_path=sha_abs_path,
                title=title,
                description=description,
                width=str(width).replace("px", "") if width else "",
                height=str(height).replace("px", "") if height else "",
                parameters_list=parameters_list,
                scripts_list=scripts_list
            )
            logging.debug("Parsed SHA %s (params=%d scripts=%d)", sha_abs_path, len(parameters_list), len(scripts_list))
            return meta
        except Exception as exc:
            logging.exception("Ошибка парсинга SHА %s: %s", sha_abs_path, exc)
            return SHAMetadata(abs_path=sha_abs_path)


# ---------- Парсер HTM ----------
class HTMParser:
    """
    Разбирает один HTM-файл, находит элементы hsc.shape.1 и собирает список HMIElement.
    """

    def __init__(self, parser: str = DEFAULT_PARSER):
        self.parser = parser

    def parse(self, htm_path: Path) -> List[HMIElement]:
        """Парсит HTM и возвращает список HMIElement."""
        logging.info("Парсинг HTM: %s", htm_path)
        try:
            # Попытки открыть с разными кодировками
            raw = None
            for enc in ("windows-1252", "windows-1251", "utf-8", "latin-1"):
                try:
                    raw = htm_path.read_text(encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            if raw is None:
                logging.error("Не удалось прочитать HTM %s", htm_path)
                return []

            soup = BeautifulSoup(raw, self.parser)

            # Ищем все объекты/дивы с классом содержащим 'hsc.shape.1'
            elements = soup.find_all(lambda tag: tag.name in ("div", "object") and tag.has_attr("class") and "hsc.shape.1" in " ".join(tag.get("class", [])))

            result: List[HMIElement] = []
            for el in elements:
                el_id = el.get("id", "N/A")
                # Берём последнее значение class
                el_class_attr = el.get("class", [])
                el_class = el_class_attr[-1] if isinstance(el_class_attr, (list, tuple)) and el_class_attr else (el.get("class") or "N/A")
                src = el.get("src", "N/A")
                link_type = el.get("linktype", "N/A")
                hdx_raw = el.get("hdxproperties", "") or ""
                params_raw = el.get("parameters", "") or ""

                parsed_hdx = parse_hdxproperties(hdx_raw)
                parsed_params = parse_parameters_attribute(params_raw)

                result.append(HMIElement(
                    htm_file=htm_path,
                    element_id=el_id,
                    element_class=str(el_class),
                    src=str(src),
                    link_type=str(link_type),
                    hdxproperties_raw=hdx_raw,
                    parameters_raw=params_raw,
                    hdx_parsed=parsed_hdx,
                    parameters_parsed=parsed_params
                ))
            logging.info("  Найдено элементов: %d", len(result))
            return result
        except Exception as e:
            logging.exception("Ошибка при парсинге HTM %s: %s", htm_path, e)
            return []


# ---------- Сканер проекта ----------
class ProjectScanner:
    """
    Сканирует корневую папку и находит HTM-файлы.
    Для каждого HTM собирает элементы и формирует список абсолютных путей SHA,
    разрешая относительные SRC относительно папки HTM.
    """

    def __init__(self, root_dir: Path, htm_parser: HTMParser):
        self.root_dir = root_dir
        self.htm_parser = htm_parser

    def find_htm_files(self) -> List[Path]:
        """Рекурсивно ищем .htm и .html файлы."""
        found = list(self.root_dir.rglob("*.htm")) + list(self.root_dir.rglob("*.html"))
        logging.info("Найдено HTM файлов в %s: %d", self.root_dir, len(found))
        return found

    def collect_elements_and_shas(self) -> (Dict[Path, List[HMIElement]], Dict[Path, Set[Path]]):
        """
        Возвращает:
         - mapping_htm_to_elements: Map[htm_path -> List[HMIElement]]
         - mapping_htm_to_sha_paths: Map[htm_path -> Set[absolute sha paths]]
        """
        mapping_htm_to_elements: Dict[Path, List[HMIElement]] = {}
        mapping_htm_to_sha_paths: Dict[Path, Set[Path]] = {}

        htm_files = self.find_htm_files()
        for htm in htm_files:
            elements = self.htm_parser.parse(htm)
            mapping_htm_to_elements[htm] = elements
            sha_paths: Set[Path] = set()
            htm_dir = htm.parent

            for el in elements:
                src = el.src or ""
                # Если src отсутствует или не .sha — пропускаем
                if not src or not isinstance(src, str):
                    continue
                # Некоторые src могут быть javascript вызовами или пустыми
                src_clean = src.replace("\\", "/").strip().strip('"').strip("'")
                if not src_clean:
                    continue
                # Если src содержит data: или http(s): - пропускаем (не локальные sha)
                if src_clean.startswith("data:") or re.match(r"^https?:/", src_clean, flags=re.I):
                    continue
                # Нормализуем путь относительно папки HTM
                sha_abs = (htm_dir / Path(src_clean)).resolve()
                sha_paths.add(sha_abs)
            mapping_htm_to_sha_paths[htm] = sha_paths
            logging.info("  HTM %s -> ссылок на SHA: %d", htm.name, len(sha_paths))

        return mapping_htm_to_elements, mapping_htm_to_sha_paths


class ExcelWriter:
    """
    Формирует Excel файл с листами:
     - General_Info
     - HMI_Tags
     - Navigation_Shapes
     - Shape_Definitions
    Заголовки формируются частично динамически (hdx keys, параметры).
    """

    def __init__(self, output_path: Path):
        self.output_path = output_path

    @staticmethod
    def _safe_value(v):
        if v is None:
            return ""
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, (list, dict)):
            try:
                return json.dumps(v, ensure_ascii=False)
            except Exception:
                return str(v)
        return str(v)

    def write(self,
              general_info: List[Dict[str, Any]],
              hmi_rows: List[Dict[str, Any]],
              nav_rows: List[Dict[str, Any]],
              shape_defs: List[Dict[str, Any]]):
        """
        Записывает все листы в XLSX.
        Заголовки для HMI_Tags и Navigation_Shapes собираются по ключам данных.
        """
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # --- General_Info ---
        sheet = wb.create_sheet("General_Info")
        general_headers = ["Файл", "Заголовок Мнемосхемы", "HMIWeb Template", "Размеры Фона (Width x Height)", "Цвет Фона"]
        sheet.append(general_headers)
        for row in general_info:
            sheet.append([self._safe_value(row.get(h, "")) for h in general_headers])

        # --- HMI_Tags ---
        if hmi_rows:
            hmi_headers = self._collect_headers(hmi_rows, base_static=[
                "Файл", "ID", "Тип Элемента (Class)", "ItemProperty1", "Itemproperty2", "Data Source", "Сервер",
                "Link Type", "hdxproperties", "Источник Формы (SRC)"
            ])
            hmi_sheet = wb.create_sheet("HMI_Tags")
            hmi_sheet.append(hmi_headers)
            for r in hmi_rows:
                hmi_sheet.append([self._safe_value(r.get(h, "")) for h in hmi_headers])

        # --- Navigation_Shapes ---
        if nav_rows:
            nav_headers = self._collect_headers(nav_rows, base_static=[
                "Файл", "ID", "Тип Элемента (Class)", "Параметры (Parameters)", "Link Type", "hdxproperties",
                "Источник Формы (SRC)"
            ])
            nav_sheet = wb.create_sheet("Navigation_Shapes")
            nav_sheet.append(nav_headers)
            for r in nav_rows:
                nav_sheet.append([self._safe_value(r.get(h, "")) for h in nav_headers])

        # --- Shape_Definitions ---
        if shape_defs:
            shape_headers = ["Файл SHAPE (SRC)", "Title", "Description", "Width", "Height",
                             "Параметры (Count)", "Параметры (Name, Type, Default)",
                             "Скрипты (Count)", "Скрипты (ID)", "Скрипты (Language)", "Скрипты (Event)",
                             "Скрипты (HasCode)"]
            sd_sheet = wb.create_sheet("Shape_Definitions")
            sd_sheet.append(shape_headers)
            for meta_row in shape_defs:
                # scripts detail может быть списком — сплющим в строки
                scripts = meta_row.get("Scripts_Detail", [])
                if scripts:
                    for s in scripts:
                        row = []
                        # основные поля
                        flat = meta_row.get("Shape_Flat", {})
                        for h in shape_headers[:7]:
                            row.append(self._safe_value(flat.get(h, "")))
                        # Скриптовые поля
                        row.append(self._safe_value(len(scripts)))
                        row.append(self._safe_value(s.get("ID", "")))
                        row.append(self._safe_value(s.get("Language", "")))
                        row.append(self._safe_value(s.get("Event", "")))
                        row.append(self._safe_value(s.get("HasCode", "")))
                        sd_sheet.append(row)
                else:
                    row = []
                    flat = meta_row.get("Shape_Flat", {})
                    for h in shape_headers[:7]:
                        row.append(self._safe_value(flat.get(h, "")))
                    row += ["", "", "", "", ""]
                    sd_sheet.append(row)

        # Попытка сохранить
        try:
            wb.save(self.output_path)
            logging.info("Результат сохранён в %s", self.output_path)
        except Exception as e:
            logging.exception("Ошибка сохранения Excel: %s", e)

    @staticmethod
    def _collect_headers(rows: List[Dict[str, Any]], base_static: List[str]) -> List[str]:
        """Собирает итоговые заголовки: base_static + все ключи из rows (в порядке)"""
        dynamic_keys: List[str] = []
        for r in rows:
            for k in r.keys():
                if k not in base_static and k not in dynamic_keys:
                    dynamic_keys.append(k)
        # Перемещаем SRC в конец, если присутствует
        headers = [h for h in base_static if h != "Источник Формы (SRC)"]
        headers += dynamic_keys
        headers += ["Источник Формы (SRC)"]
        return headers


def main(root_dir: str, output_filename: str, max_workers: int = 8):
    root = Path(root_dir).resolve()
    logging.info("Запуск. Корень: %s", root)

    htm_parser = HTMParser()
    scanner = ProjectScanner(root, htm_parser)

    # 1) Собираем элементы и список всех sha, на которые есть ссылки (по HTM)
    mapping_htm_to_elements, mapping_htm_to_sha_paths = scanner.collect_elements_and_shas()

    # 2) ПАРАЛЛЕЛЬНО парсим все уникальные SHA пути (в пределах абсолютных путей)
    # ВАЖНО: вы просили анализировать повторы: если два HTM ссылаются на разные абсолютные пути — оба будут разобраны.
    unique_sha_paths: Set[Path] = set()
    for sha_set in mapping_htm_to_sha_paths.values():
        unique_sha_paths.update(sha_set)

    logging.info("Уникальных абсолютных SHA-путей для разбора: %d", len(unique_sha_paths))

    sha_parser = SHAParser()
    sha_meta_by_path: Dict[Path, Optional[SHAMetadata]] = {}

    # Парсим SHA параллельно, но аккуратно — файл может не существовать
    def parse_one(sha_path: Path) -> (Path, Optional[SHAMetadata]):
        if not sha_path.exists():
            logging.warning("Ссылка на SHA отсутствует на диске: %s", sha_path)
            return sha_path, None
        meta = sha_parser.parse(sha_path)
        return sha_path, meta

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as exc:
        futures = {exc.submit(parse_one, p): p for p in unique_sha_paths}
        for fut in concurrent.futures.as_completed(futures):
            p = futures[fut]
            try:
                sha_path, meta = fut.result()
                sha_meta_by_path[sha_path] = meta
            except Exception as e:
                logging.exception("Ошибка при парсинге SHA %s: %s", p, e)
                sha_meta_by_path[p] = None

    # 3) Формируем результирующие таблицы
    general_info_rows = []
    hmi_tags_rows = []
    nav_rows = []
    shape_definitions_rows = []

    # Общая информация (из HTM <title>, HMIWebTemplate, body style)
    for htm_path, elements in mapping_htm_to_elements.items():
        # Общая информация: парсим файл для title + meta HMIWebTemplateDescription + body style
        try:
            raw = None
            for enc in ("windows-1252", "windows-1251", "utf-8", "latin-1"):
                try:
                    raw = htm_path.read_text(encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            if raw is None:
                continue
            soup = BeautifulSoup(raw, DEFAULT_PARSER)
            title_tag = soup.find("title")
            title = title_tag.text.strip() if title_tag else htm_path.name
            hmi_template_tag = soup.find("meta", {"name": "HMIWebTemplateDescription"})
            hmi_template = hmi_template_tag.get("content", "") if hmi_template_tag else ""
            body = soup.find("body")
            style = body.get("style", "") if body else ""
            width_m = re.search(r"width:\s*(\d+)px", style)
            height_m = re.search(r"height:\s*(\d+)px", style)
            bg_m = re.search(r"background-color:\s*([^;]+)", style)
            dims = ""
            if width_m and height_m:
                dims = f"{width_m.group(1)}x{height_m.group(1)}"
            bg_color = bg_m.group(1).strip() if bg_m else "white"
            general_info_rows.append({
                "Файл": str(htm_path.name),
                "Заголовок Мнемосхемы": title,
                "HMIWeb Template": hmi_template,
                "Размеры Фона (Width x Height)": dims,
                "Цвет Фона": bg_color
            })
        except Exception:
            logging.exception("Ошибка при извлечении General_Info из %s", htm_path)

    # Для каждого HTM -> элемента формируем строки в HMI_Tags или Navigation_Shapes
    for htm_path, elements in mapping_htm_to_elements.items():
        for el in elements:
            # базовые поля
            base = {
                "Файл": str(htm_path),  # полный путь, ExcelWriter может менять/обрабатывать
                "ID": el.element_id,
                "Тип Элемента (Class)": el.element_class,
                "Источник Формы (SRC)": el.src,
                "Link Type": el.link_type,
                "hdxproperties": el.hdxproperties_raw,
            }
            # добавляем динамические hdx поля
            for k, v in el.hdx_parsed.items():
                base[f"hdxproperties - {k}"] = v

            # если среди параметров есть ItemProperty1 — считаем что это HMI_Tags
            params = el.parameters_parsed or {}
            if "ItemProperty1" in params:
                # HMI_Tags: включаем статические поля ItemProperty1, Itemproperty2, Data Source, Сервер
                row = dict(base)
                row.update({
                    "ItemProperty1": params.get("ItemProperty1", ""),
                    "Itemproperty2": params.get("Itemproperty2", ""),
                    "Data Source": params.get("Data Source", ""),
                    "Сервер": params.get("Server Control", "")
                })
                # добавим sha-merge поля если есть соответствующий sha
                # разрешим абс путь
                src = el.src.strip()
                if src and src.lower().endswith(".sha"):
                    sha_abs = (htm_path.parent / src).resolve()
                    sha_meta = sha_meta_by_path.get(sha_abs)
                    if sha_meta:
                        # добавляем non-script поля
                        row["Title"] = sha_meta.title
                        row["Description"] = sha_meta.description
                        row["Width"] = sha_meta.width or ""
                        row["Height"] = sha_meta.height or ""
                        row["Параметры (Count)"] = len(sha_meta.parameters_list)
                        row["Параметры (Name, Type, Default)"] = json.dumps(sha_meta.parameters_list, ensure_ascii=False)
                        row["Скрипты (Count)"] = len(sha_meta.scripts_list)
                        # добавляем первые скриптовые поля (скрипты конкретно добавляются в merge step в Excel)
                        # Но оставим отдельное поле для удобства
                    else:
                        # sha отсутствует или не разобран
                        row["Title"] = ""
                        row["Description"] = ""
                        row["Width"] = ""
                        row["Height"] = ""
                        row["Параметры (Count)"] = ""
                        row["Параметры (Name, Type, Default)"] = ""
                        row["Скрипты (Count)"] = ""
                hmi_tags_rows.append(row)
            else:
                # Navigation_Shapes
                row = dict(base)
                row["Параметры (Parameters)"] = el.parameters_raw
                # добавляем parsed params prefixed
                for k, v in params.items():
                    row[f"Параметры - {k}"] = v
                # добавляем sha merge fields если есть
                src = el.src.strip()
                if src and src.lower().endswith(".sha"):
                    sha_abs = (htm_path.parent / src).resolve()
                    sha_meta = sha_meta_by_path.get(sha_abs)
                    if sha_meta:
                        row["Title"] = sha_meta.title
                        row["Description"] = sha_meta.description
                        row["Width"] = sha_meta.width or ""
                        row["Height"] = sha_meta.height or ""
                        row["Параметры (Count)"] = len(sha_meta.parameters_list)
                        row["Параметры (Name, Type, Default)"] = json.dumps(sha_meta.parameters_list, ensure_ascii=False)
                        row["Скрипты (Count)"] = len(sha_meta.scripts_list)
                    else:
                        row["Title"] = ""
                        row["Description"] = ""
                        row["Width"] = ""
                        row["Height"] = ""
                        row["Параметры (Count)"] = ""
                        row["Параметры (Name, Type, Default)"] = ""
                        row["Скрипты (Count)"] = ""
                # добавить hdx парсинг
                for k, v in el.hdx_parsed.items():
                    row[f"hdxproperties - {k}"] = v
                nav_rows.append(row)

    # --- Shape_Definitions: формируем список словарей, содержащих Shape_Flat и Scripts_Detail
    for sha_path, meta in sha_meta_by_path.items():
        if meta is None:
            continue
        shape_row = {
            "Shape_Flat": meta.to_flat_dict(),
            "Scripts_Detail": meta.scripts_list,
        }
        shape_definitions_rows.append(shape_row)

    # 4) Запись в XLSX
    writer = ExcelWriter(Path(output_filename))
    writer.write(
        general_info=general_info_rows,
        hmi_rows=hmi_tags_rows,
        nav_rows=nav_rows,
        shape_defs=shape_definitions_rows
    )

    logging.info("Готово.")


# ---------- CLI ----------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Парсер HTM+SHA -> XLSX")
    parser.add_argument("--root", "-r", help="Корневая папка с мнемосхемами (по умолчанию текущая)", default=".")
    parser.add_argument("--output", "-o", help="Имя выходного файла Excel", default="Result_improved.xlsx")
    parser.add_argument("--workers", "-w", help="Число потоков для разбора SHA", type=int, default=8)
    args = parser.parse_args()

    main(root_dir=args.root, output_filename=args.output, max_workers=args.workers)
