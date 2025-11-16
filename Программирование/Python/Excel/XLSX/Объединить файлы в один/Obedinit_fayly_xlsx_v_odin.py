import pandas as pd
import glob
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Подавляем предупреждения openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

output_filename = "output.xlsx"
sheet_name = "Models"

all_files = [f for f in glob.glob("*.xlsx") if f != output_filename]

if not all_files:
    print("Нет файлов для обработки.")
else:
    print(f"Обнаружены файлы для обработки ({len(all_files)} шт.):")

    try:
        # Загружаем первый файл как основу
        first_file = all_files[0]
        wb = load_workbook(first_file)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"В первом файле {first_file} отсутствует лист '{sheet_name}'")

        ws = wb[sheet_name]

        # Собираем данные из всех файлов (кроме первого)
        data_frames = []

        # Сначала добавляем данные из самого первого файла
        first_df = pd.read_excel(first_file, sheet_name=sheet_name)
        data_frames.append(first_df)
        print(f"✓ {first_file} - лист '{sheet_name}' ({len(first_df)} строк)")

        # Затем добавляем данные из остальных файлов
        for file in all_files[1:]:
            try:
                df = pd.read_excel(file, sheet_name=sheet_name)
                data_frames.append(df)
                print(f"✓ {file} - лист '{sheet_name}' ({len(df)} строк)")
            except Exception as e:
                print(f"⚠ {file} - ошибка: {str(e)}")
                continue

        if len(data_frames) < 2:
            raise ValueError("Нет дополнительных данных для объединения")

        # Объединяем данные
        combined_df = pd.concat(data_frames, ignore_index=True)

        if not all(col in combined_df.columns for col in ["Type", "Path"]):
            raise ValueError("Отсутствуют столбцы 'Type' или 'Path'")

        # Сортируем
        result_df = combined_df.sort_values(by=["Type", "Path"], ignore_index=True)

        # Очищаем существующий лист (кроме заголовков)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Добавляем новые данные в существующий лист
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

        # Сохраняем результат
        wb.save(output_filename)
        print(f"\nУспешно! Результат сохранён в {output_filename}")
        print(f"Обработано файлов: {len(data_frames)}")
        print(f"Всего строк: {len(result_df)}")

    except Exception as e:
        print(f"\nОшибка при обработке: {str(e)}")

