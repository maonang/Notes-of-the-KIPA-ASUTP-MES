import pandas as pd
import glob
import warnings
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
from datetime import datetime

# Подавляем предупреждения openpyxl, которые могут возникать при работе с файлами
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

OUTPUT_FILENAME = "output.xlsx"
LOG_FILENAME = "log.log"  # Имя файла для сохранения логов
PROGRESS_STEP = 5000  # Шаг для вывода прогресса в консоль (каждые N строк)
# Список заголовков столбцов, который будет автоматически определен из первого файла.
COLUMN_HEADERS = []
# Словарь для хранения стилей ячеек из второй строки первого файла
CELL_STYLES = {}


def log_message(message, log_to_file=True):
    """
    Выводит сообщение в консоль и записывает его в log.log с временной меткой.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(message)
    if log_to_file:
        try:
            with open(LOG_FILENAME, "a", encoding="utf-8") as f:
                f.write(log_line + "\n")
        except Exception as e:
            print(f"Ошибка: Не удалось записать в {LOG_FILENAME}. {str(e)}")


def get_first_sheet_name(filepath):
    """
    Возвращает имя первого (активного) листа в файле Excel.
    """
    try:
        xls = pd.ExcelFile(filepath)
        return xls.sheet_names[0] if xls.sheet_names else None
    except Exception as e:
        log_message(f"Ошибка при чтении имени листа из {filepath}: {str(e)}")
        return None


# Очищаем лог-файл перед началом работы
try:
    with open(LOG_FILENAME, "w", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Начало обработки\n")
except:
    pass

# Ищем все файлы .xlsx в текущей директории, исключая выходной файл
all_files = [f for f in glob.glob("*.xlsx") if f != OUTPUT_FILENAME]

if not all_files:
    log_message("Предупреждение: Нет файлов для обработки (.xlsx).")
else:
    log_message(f"Уведомление: Обнаружены файлы для обработки ({len(all_files)} шт.):")

    # Определяем максимальную длину имени файла для красивого выравнивания
    max_file_name_len = max(len(f) for f in all_files)

    try:
        # 1. Автоматически определяем имя листа и заголовки столбцов
        first_file = all_files[0]

        # Получаем имя первого листа
        sheet_name = get_first_sheet_name(first_file)
        if not sheet_name:
            raise ValueError(f"Предупреждение: Не удалось определить имя листа в первом файле: {first_file}")

        log_message(f"Уведомление: Имя листа для обработки: '{sheet_name}'")

        # Читаем данные из первого файла для получения списка столбцов и их порядка
        log_message(f"Уведомление: Чтение эталонного файла {first_file} для определения структуры...")
        first_df = pd.read_excel(first_file, sheet_name=sheet_name)

        # Обновляем список заголовков столбцов, сохраняя их исходный порядок
        COLUMN_HEADERS = list(first_df.columns)
        if not COLUMN_HEADERS:
            raise ValueError(f"Предупреждение: Не найдены заголовки столбцов в файле: {first_file}")

        log_message(f"Уведомление: Ожидаемый порядок столбцов ({len(COLUMN_HEADERS)} шт.): {', '.join(COLUMN_HEADERS)}")

        # 2. Собираем данные из всех файлов
        data_frames = []

        # Добавляем данные из самого первого файла и форматируем вывод
        data_frames.append(first_df)
        log_message(f"{first_file:<{max_file_name_len}} - лист '{sheet_name}' ({len(first_df):>7} строк)")

        # Добавляем данные из остальных файлов
        for file in all_files[1:]:
            try:
                log_message(f"Уведомление: Чтение файла {file}...")
                df = pd.read_excel(file, sheet_name=sheet_name)

                # Проверка наличия всех заголовков в требуемом порядке
                if list(df.columns) != COLUMN_HEADERS:
                    log_message(
                        f"Предупреждение: {file:<{max_file_name_len}} - Порядок или состав столбцов не совпадает. Пропуск.")
                    continue

                data_frames.append(df)
                # Форматируем вывод
                log_message(f"{file:<{max_file_name_len}} - лист '{sheet_name}' ({len(df):>7} строк)")
            except Exception as e:
                log_message(f"Предупреждение: {file:<{max_file_name_len}} - ошибка при чтении: {str(e)}. Пропуск.")
                continue

        if len(data_frames) == 0:
            raise ValueError("Уведомление: Не удалось собрать данные ни из одного файла для объединения.")

        # 3. Объединяем данные, сохраняя исходный порядок строк и столбцов
        log_message(f"Уведомление: Объединяю данные из {len(data_frames)} датафреймов...")
        combined_df = pd.concat(data_frames, ignore_index=True)
        result_df = combined_df[COLUMN_HEADERS]
        total_rows = len(result_df)
        log_message(f"Уведомление: Объединение завершено. Всего строк для записи: {total_rows}.")

        # 4. Сохранение результата в новый/существующий файл с сохранением форматирования
        log_message(f"Уведомление: Загружаю {first_file} для копирования стилей и форматирования...")

        # Загружаем первый файл openpyxl (для получения стилей)
        try:
            wb = load_workbook(first_file)
        except Exception as e:
            # Создаем новую книгу, если загрузка не удалась, то используем только стандартные стили
            wb = Workbook()
            if 'Sheet' in wb.sheetnames and not wb['Sheet'].cell(1, 1).value:
                del wb['Sheet']
            log_message(
                f"Уведомление: Не удалось загрузить {first_file} для копирования стилей. Будут применены стандартные границы. Ошибка: {str(e)}")

        # Если лист с нужным именем не существует, создаем его
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # Получение стилей из второй строки (первая строка данных) ---
        if ws.max_row >= 2:
            log_message(f"Уведомление: Копирование стилей из второй строки первого файла...")
            # Сохраняем стили всех ячеек второй строки
            for c_idx in range(1, len(COLUMN_HEADERS) + 1):
                cell_to_copy = ws.cell(row=2, column=c_idx)

                # Копируем стили: шрифт, заполнение, выравнивание, формат чисел, защита, граница
                CELL_STYLES[c_idx] = {
                    'font': copy(cell_to_copy.font),
                    'fill': copy(cell_to_copy.fill),
                    'alignment': copy(cell_to_copy.alignment),
                    'number_format': cell_to_copy.number_format,
                    'protection': copy(cell_to_copy.protection),
                    'border': copy(cell_to_copy.border)
                }
            log_message(f"Уведомление: Стили успешно скопированы для {len(CELL_STYLES)} столбцов.")
        else:
            log_message(
                "Предупреждение: В первом файле нет второй строки данных для копирования стилей. Будут применены стандартные границы.")

        # Очищаем существующий лист (кроме заголовков)
        if ws.max_row > 1:
            log_message(f"Уведомление: Очистка существующих данных в листе '{sheet_name}'...")
            ws.delete_rows(2, ws.max_row)

        # Заново добавляем заголовки в нужном порядке
        for c_idx, col_name in enumerate(COLUMN_HEADERS, 1):
            ws.cell(row=1, column=c_idx, value=col_name)

        log_message(f"Уведомление: Вставка {total_rows} строк данных...")
        log_message(f"Уведомление: Цикл вставки данных запущен. Ожидайте обновления прогресса в консоли.",
                    log_to_file=True)

        # Добавляем новые данные
        # Стандартные тонкие границы (используются, если CELL_STYLES пусты)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Начинаем вставку со строки 2 (после заголовков)
        for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=False), 2):
            current_row_num = r_idx - 1  # Номер строки в датафрейме (от 1 до total_rows)

            if current_row_num == round(total_rows * 0.5):
                log_message(f"Уведомление: Достигнута середина вставки данных ({current_row_num} строк)",
                            log_to_file=True)

            if total_rows > PROGRESS_STEP and current_row_num > 0 and current_row_num % PROGRESS_STEP == 0:
                print(f"Обработка строк: {current_row_num}/{total_rows} ({current_row_num / total_rows:.1%})", end='\r')

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.value = value
                if c_idx in CELL_STYLES:
                    # Применяем скопированные стили
                    styles = CELL_STYLES[c_idx]
                    cell.font = styles['font']
                    cell.fill = styles['fill']
                    cell.alignment = styles['alignment']
                    cell.number_format = styles['number_format']
                    cell.protection = styles['protection']
                    cell.border = styles['border']
                else:
                    # Если стили не были скопированы, применяем базовые границы
                    cell.border = thin_border

        # Очищаем строку прогресса в консоли и добавляем новую строку
        print(" " * 80, end='\r')
        sys.stdout.flush()
        log_message(f"Уведомление: Запись всех {total_rows} строк в лист завершена.")
      
        # Сохраняем результат
        log_message(f"Уведомление: Сохраняю итоговый файл {OUTPUT_FILENAME}. Это может занять время...")
        wb.save(OUTPUT_FILENAME)

        # 5. Вывод итоговой информации
        log_message(f"\nУведомление: Процесс завершен!")
        log_message(f"Уведомление: Результат сохранён в {OUTPUT_FILENAME}")
        log_message(f"Уведомление: Обработано файлов: {len(data_frames)}")
        log_message(f"Уведомление: Всего строк в результате: {len(result_df)}")

    except Exception as e:
        log_message(f"\nОбработка завершена с ошибкой: {str(e)}", log_to_file=True)

# Завершающая запись в лог
log_message(f"Обработка завершена", log_to_file=True)
