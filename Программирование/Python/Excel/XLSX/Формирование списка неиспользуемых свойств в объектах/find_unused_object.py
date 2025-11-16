import json
import openpyxl
import os
import pandas as pd
import re
import shutil
import warnings

# Отключение предупреждений UserWarning
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def normalize_path(path):
    if isinstance(path, str):
        return re.sub(r'\\\\+', r'\\', path)
    return path

def find_unused_id(file_path: str) -> list:
    df = pd.read_excel(file_path)
    df['Path'] = df['Path'].fillna('')
    set_Path_Property_Instance = set(normalize_path(path) for path in df[df['PropertyType'] == 'Instance']['Path'].dropna())

    filtered_df = df[df['Type'] == 'PropertyConfiguration']
    all_paths = set(normalize_path(path) for path in filtered_df['Path'])
    used_path = set()
    used_path_calc = set()

    for _, row in filtered_df.iterrows():
        configuration = row['Configuration']
        path = row['Path']
        if pd.notna(configuration):
            try:
                config_data = json.loads(configuration)

                variables_used = False
                expression_used = False

                if 'variables' in config_data:
                    for variable in config_data['variables']:
                        if 'value' in variable and variable['value'].strip():
                            variables_used = True
                            used_path.add(variable['value'])
                if 'expression' in config_data and config_data['expression'].strip():
                    expression_used = True
                if variables_used and expression_used:
                    used_path_calc.add(path)
            except json.JSONDecodeError as e:
                print(f"Ошибка декодирования JSON для Id={row['Id']}: {e}")

    used_path = set(normalize_path(path) for path in used_path)

    print(f"Общее количество путей до всех свойств в поле Path: {len(all_paths)}")
    print(f"Общее количество путей до свойств с формулами в поле Path: {len(used_path_calc)}")
    print(f"Общее количество путей до всех свойств в поле Configuration: {len(used_path)}")
    print(f"Общее количество путей с значением Instance в поле PropertyType: {len(set_Path_Property_Instance)}")

    unused_paths = all_paths - used_path - set_Path_Property_Instance - used_path_calc
    print(f"Количество неиспользуемых путей: {len(unused_paths)}")
    return list(unused_paths)

def delete_rows_by_id(file_path, list_id):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Models']

    rows_to_delete = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=4).value
        if cell_value not in list_id:
            rows_to_delete.append(row)
    # Удаляем строки в обратном порядке, чтобы не нарушить индексацию
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    wb.save(file_path)
    print(f"Файл '{file_path}' успешно обновлен.")

print('Программа предназначена для формирования списка неиспользуемых свойств в объектах.\n'
      'Признаки для удаления записи:\n'
      '\t1) Значение поля Type не равно Property или PropertyConfiguration;\n'
      '\t2) Значение поля PropertyType равно Instance;\n'
      '\t3) Свойство не встречается в поле Configuration.\n')

while True:
    s_file_name_input = input("Введите название xlsx файла: ")
    if not s_file_name_input.endswith('.xlsx'):
        s_file_name_input += '.xlsx'
    if not os.path.exists(s_file_name_input):
        continue
    else:
        # Формирование имени выходного файла
        file_name, file_format = os.path.splitext(s_file_name_input)
        s_file_name_output = f"{file_name}_copy{file_format}"
        if os.path.exists(s_file_name_output):
            os.remove(s_file_name_output)
        shutil.copy(s_file_name_input, s_file_name_output)

        try:
            list_id = find_unused_id(s_file_name_input)
            # Удаление строк в выходном файле
            delete_rows_by_id(s_file_name_output, list_id)
        except Exception as e:
            print(f"Произошла ошибка: {e}")
        print()

