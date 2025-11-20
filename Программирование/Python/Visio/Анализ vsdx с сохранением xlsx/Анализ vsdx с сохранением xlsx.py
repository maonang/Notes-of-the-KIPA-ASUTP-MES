from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
from vsdx import VisioFile
import os
import pandas as pd
import re
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def clean_text(text):
    """Очистка и нормализация текста с обработкой кодировки"""
    if text is None:
        return None

    # Преобразуем в строку если это не строка
    if not isinstance(text, str):
        text = str(text)

    # Удаляем лишние пробелы и переносы
    text = text.strip()

    # Заменяем проблемные символы кодировки
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', text)  # Удаляем управляющие символы
    text = re.sub(r'\s+', ' ', text)  # Заменяем множественные пробелы на один

    # Декодируем распространенные проблемы с кодировкой
    encoding_problems = {
        'â€': '"',
        'â€“': '-',
        'â€”': '—',
        'â€¢': '•',
        'â„¢': '™',
        'â€¦': '…',
        'â€˜': "'",
        'â€™': "'",
        'â€œ': '"',
        'â€': '"',
    }

    for problem, replacement in encoding_problems.items():
        text = text.replace(problem, replacement)

    return text if text else None


def get_connector_properties(shape):
    """Извлекает специальные свойства соединителей"""
    connector_props = {}

    try:
        # Проверяем, является ли фигура соединителем
        if hasattr(shape, 'xml') and shape.xml is not None:
            # Ищем атрибуты соединителей
            connector_props['is_connector'] = True

            # Получаем начальную и конечную точки соединителя
            if hasattr(shape, 'begin_x') and hasattr(shape, 'begin_y'):
                connector_props['begin_x'] = shape.begin_x
                connector_props['begin_y'] = shape.begin_y
                connector_props['end_x'] = shape.end_x
                connector_props['end_y'] = shape.end_y

            # Получаем связанные фигуры (если доступно)
            if hasattr(shape, 'connects'):
                connector_props['connected_shapes'] = []
                for connect in shape.connects:
                    try:
                        connector_props['connected_shapes'].append({
                            'shape_id': getattr(connect, 'shape_id', None),
                            'from_cell': getattr(connect, 'from_cell', None),
                            'to_cell': getattr(connect, 'to_cell', None)
                        })
                    except Exception:
                        pass

            # Определяем тип соединителя
            shape_xml = shape.xml
            if shape_xml.get('Master'):
                master = shape_xml.get('Master')
                if 'Dynamic' in str(master):
                    connector_props['connector_type'] = 'Dynamic'
                elif 'Straight' in str(master):
                    connector_props['connector_type'] = 'Straight'
                elif 'Curved' in str(master):
                    connector_props['connector_type'] = 'Curved'
                else:
                    connector_props['connector_type'] = 'Other'
            else:
                connector_props['connector_type'] = 'Unknown'

    except Exception as e:
        logger.debug(f"Ошибка при извлечении свойств соединителя: {e}")
        connector_props['is_connector'] = False

    return connector_props


def get_shape_properties(shape):
    """Безопасно извлекает свойства фигуры"""
    props = {}

    # Базовые свойства
    safe_attrs = ['ID', 'x', 'y', 'width', 'height', 'text',
                  'fill_color', 'line_color', 'line_weight',
                  'text_color', 'font', 'name', 'rotation']

    for attr in safe_attrs:
        try:
            value = getattr(shape, attr, None)
            # Обрабатываем текст
            if attr == 'text' and value:
                value = clean_text(value)
            props[attr] = value
        except Exception as e:
            logger.debug(f"Ошибка получения атрибута {attr}: {e}")
            props[attr] = None

    # Дополнительные свойства из XML
    try:
        props['shape_name'] = shape.master_name if hasattr(shape, 'master_name') else None
        props['hyperlinks'] = shape.hyperlinks if hasattr(shape, 'hyperlinks') else []
        props['shape_type'] = shape.xml.get('Type') if hasattr(shape, 'xml') else None
        props['layer'] = shape.layer_name if hasattr(shape, 'layer_name') else None

        # Обрабатываем гиперссылки (очищаем текст)
        if props['hyperlinks']:
            cleaned_hyperlinks = []
            for link in props['hyperlinks']:
                try:
                    if hasattr(link, 'description') and link.description:
                        link.description = clean_text(link.description)
                    if hasattr(link, 'address') and link.address:
                        link.address = clean_text(link.address)
                    cleaned_hyperlinks.append(link)
                except Exception:
                    cleaned_hyperlinks.append(link)
            props['hyperlinks'] = cleaned_hyperlinks

    except Exception as e:
        logger.debug(f"Ошибка получения дополнительных свойств: {e}")

    # Свойства соединителей
    connector_props = get_connector_properties(shape)
    props.update(connector_props)

    return props


def extract_shapes_recursive(shape, page_name, elements):
    """Рекурсивно извлекает вложенные фигуры"""
    try:
        element = {
            'Page': page_name,
            **get_shape_properties(shape)
        }
        elements.append(element)

        # Обработка дочерних фигур
        if hasattr(shape, 'child_shapes'):
            for child in shape.child_shapes:
                extract_shapes_recursive(child, page_name, elements)
    except Exception as e:
        logger.warning(f"Ошибка при рекурсивном извлечении фигуры: {e}")


def extract_all_elements(vsdx_path: str) -> list[dict]:
    """Извлекает все элементы и их параметры из .vsdx"""
    elements = []

    try:
        with VisioFile(vsdx_path) as vis:
            for page in vis.pages:
                logger.info(f"Обработка страницы: {page.name}")

                for shape in page.child_shapes:
                    try:
                        extract_shapes_recursive(shape, page.name, elements)
                    except Exception as e:
                        logger.warning(f"Ошибка при обработке фигуры на странице {page.name}: {e}")
                        # Пытаемся извлечь хотя бы базовую информацию
                        try:
                            element = {
                                'Page': page.name,
                                'ID': getattr(shape, 'ID', 'Unknown'),
                                'name': getattr(shape, 'name', 'Error'),
                                'text': 'Ошибка извлечения',
                                'error': str(e)
                            }
                            elements.append(element)
                        except:
                            pass

    except Exception as e:
        logger.error(f"Критическая ошибка при обработке файла {vsdx_path}: {e}")
        # Добавляем запись об ошибке
        elements.append({
            'Page': 'Error',
            'ID': 'N/A',
            'name': 'File Error',
            'text': f'Ошибка обработки файла: {str(e)}',
            'error': str(e)
        })

    return elements


def apply_color_formatting(ws, color_column_index):
    """Добавляет цветовую заливку для ячеек с цветами"""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=color_column_index, max_col=color_column_index):
        for cell in row:
            if cell.value:
                # Пытаемся извлечь цвет из значения (формат "#RRGGBB" или "RGB(R,G,B)")
                color = None
                if isinstance(cell.value, str):
                    # Для формата "#RRGGBB"
                    hex_match = re.match(r'^#([A-Fa-f0-9]{6})$', cell.value)
                    if hex_match:
                        color = hex_match.group(1)
                    else:
                        # Для формата "RGB(R,G,B)"
                        rgb_match = re.match(r'^RGB\((\d+),\s*(\d+),\s*(\d+)\)$', cell.value)
                        if rgb_match:
                            r, g, b = rgb_match.groups()
                            color = f"{int(r):02X}{int(g):02X}{int(b):02X}"

                if color:
                    try:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        # Делаем текст белым для темных цветов
                        if is_dark_color(color):
                            cell.font = Font(color="FFFFFF")
                    except Exception:
                        # Если цвет невалидный, оставляем без изменений
                        pass


def is_dark_color(hex_color):
    """Определяет, является ли цвет темным (для выбора цвета текста)"""
    if not hex_color or len(hex_color) != 6:
        return False
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        brightness = (r * 299 + g * 587 + b * 114) / 1000
        return brightness < 128
    except:
        return False


def save_to_excel(data: list[dict], output_path: str):
    """Сохраняет данные в Excel файл с форматированием"""
    if not data:
        logger.warning("Нет данных для сохранения")
        return

    # Очищаем данные перед сохранением в DataFrame
    cleaned_data = []
    for item in data:
        cleaned_item = {}
        for key, value in item.items():
            # Очищаем текстовые значения
            if isinstance(value, str):
                cleaned_item[key] = clean_text(value)
            elif isinstance(value, list):
                # Обрабатываем списки (например, гиперссылки)
                cleaned_list = []
                for list_item in value:
                    if isinstance(list_item, str):
                        cleaned_list.append(clean_text(list_item))
                    else:
                        cleaned_list.append(list_item)
                cleaned_item[key] = str(cleaned_list)  # Преобразуем в строку для Excel
            else:
                cleaned_item[key] = value
        cleaned_data.append(cleaned_item)

    try:
        df = pd.DataFrame(cleaned_data)

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Visio Elements"

        # Заголовки
        headers = list(cleaned_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        # Данные
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        # Авто-ширина колонок
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)  # Ограничиваем максимальную ширину
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Применяем цветовое форматирование
        color_columns = ['fill_color', 'line_color', 'text_color']
        for color_col in color_columns:
            if color_col in headers:
                color_col_index = headers.index(color_col) + 1
                apply_color_formatting(ws, color_col_index)

        wb.save(output_path)
        logger.info(f"Файл успешно сохранен: {output_path}")

    except Exception as e:
        logger.error(f"Ошибка при сохранении в Excel: {e}")
        raise


if __name__ == '__main__':
    input_dir = os.getcwd()
    vsdx_files = [f for f in os.listdir(input_dir) if f.endswith(".vsdx")]

    if not vsdx_files:
        logger.info("Файлы .vsdx не найдены в текущей директории")
    else:
        logger.info(f"Найдено файлов .vsdx: {len(vsdx_files)}")

        for file in tqdm(vsdx_files, desc="Обработка VSDX файлов"):
            logger.info(f"Обработка файла: {file}")
            try:
                elements_data = extract_all_elements(file)

                if elements_data:
                    output_file = file.replace(".vsdx", "_elements.xlsx")
                    save_to_excel(elements_data, output_file)
                    logger.info(f"Данные сохранены в: {output_file}")
                    logger.info(f"Извлечено элементов: {len(elements_data)}")
                else:
                    logger.warning(f"Не удалось извлечь данные из файла: {file}")

            except Exception as e:
                logger.error(f"Ошибка при обработке файла {file}: {str(e)}")

    logger.info("Обработка завершена.")