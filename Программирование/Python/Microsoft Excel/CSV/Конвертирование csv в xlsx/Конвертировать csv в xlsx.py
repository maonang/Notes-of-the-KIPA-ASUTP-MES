#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import csv
import chardet
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Iterator

# Константы
EXCEL_MAX_ROWS: int = 1_048_576
CHUNK_SIZE: int = 100_000  # строк на один чанк
ENCODING_TRIAL_ORDER: List[str] = ["utf-8", "utf-8-sig", "cp1251"]
DELIMITERS: List[str] = [",", ";", "\t", "|"]
SAMPLE_SIZE: int = 50_000  # размер выборки для анализа


class CSVProcessor:
    """Класс для обработки CSV файлов и конвертации в XLSX"""

    def __init__(self):
        self.stats: Dict[str, Any] = {"processed_files": 0, "failed_files": []}

    def detect_encoding(self, path: Path) -> str:
        """
        Определяет кодировку файла.

        Args:
            path: Путь к файлу

        Returns:
            Строка с названием кодировки
        """
        raw_data = path.read_bytes()[:1_000_000]
        detection_result = chardet.detect(raw_data)
        detected_encoding = detection_result.get("encoding")

        if not detected_encoding:
            return "utf-8"

        # Пробуем определить кодировку перебором
        for encoding in ENCODING_TRIAL_ORDER:
            try:
                raw_data.decode(encoding)
                return encoding
            except (UnicodeDecodeError, Exception):
                continue

        return detected_encoding.lower()

    def detect_delimiter(self, path: Path, encoding: str) -> str:
        """
        Определяет разделитель в CSV файле.

        Args:
            path: Путь к файлу
            encoding: Кодировка файла

        Returns:
            Символ-разделитель
        """
        with open(path, "r", encoding=encoding, errors="replace") as file:
            sample_text = file.read(SAMPLE_SIZE)

        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample_text, delimiters=DELIMITERS)
            return dialect.delimiter
        except Exception:
            # Резервная эвристика на основе подсчета символов
            delimiter_counts = {
                ";": sample_text.count(";"),
                "\t": sample_text.count("\t"),
                ",": sample_text.count(",")
            }
            return max(delimiter_counts, key=delimiter_counts.get)  # type: ignore

    def estimate_total_rows(self, path: Path) -> int:
        """
        Оценивает общее количество строк в файле.

        Args:
            path: Путь к файлу

        Returns:
            Оценочное количество строк
        """
        row_count = 0
        with path.open("rb") as file:
            for _ in file:
                row_count += 1
        return row_count

    def apply_autofilter(self, xlsx_path: Path) -> None:
        """
        Применяет автофильтр ко всем листам Excel файла.

        Args:
            xlsx_path: Путь к XLSX файлу
        """
        try:
            workbook = load_workbook(xlsx_path)
            for worksheet in workbook.worksheets:
                if worksheet.max_column > 0:
                    last_column = get_column_letter(worksheet.max_column)
                    worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"
            workbook.save(xlsx_path)
        except Exception as error:
            tqdm.write(f"[WARN] Не удалось применить автофильтр: {error}")

    def process_small_csv_file(self, csv_path: Path, encoding: str, delimiter: str) -> bool:
        """
        Обрабатывает небольшие CSV файлы (до 300k строк).

        Args:
            csv_path: Путь к CSV файлу
            encoding: Кодировка файла
            delimiter: Разделитель

        Returns:
            True если обработка успешна, False в противном случае
        """
        try:
            dataframe = pd.read_csv(
                csv_path,
                sep=delimiter,
                encoding=encoding,
                dtype=str,
                keep_default_na=False
            )

            xlsx_path = csv_path.with_suffix(".xlsx")

            with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as excel_writer:
                dataframe.to_excel(excel_writer, index=False, sheet_name="Sheet1")

                # Применяем автофильтр
                worksheet = excel_writer.sheets["Sheet1"]
                last_column_index = dataframe.shape[1]
                worksheet.autofilter(0, 0, dataframe.shape[0], last_column_index - 1)

            tqdm.write(f"[INFO] Количество записей: {len(dataframe):,}")
            tqdm.write(f"[INFO] Сохранён: {xlsx_path.name}")
            return True

        except Exception as error:
            tqdm.write(f"[ERROR] Ошибка при обработке малого файла: {error}")
            return False

    def process_large_csv_file(self, csv_path: Path, encoding: str, delimiter: str, estimated_rows: int) -> bool:
        """
        Обрабатывает большие CSV файлы (свыше 300k строк) с разбивкой на чанки.

        Args:
            csv_path: Путь к CSV файлу
            encoding: Кодировка файла
            delimiter: Разделитель
            estimated_rows: Оценочное количество строк

        Returns:
            True если обработка успешна, False в противном случае
        """
        try:
            chunk_iterator = pd.read_csv(
                csv_path,
                sep=delimiter,
                encoding=encoding,
                dtype=str,
                chunksize=CHUNK_SIZE,
                keep_default_na=False,
                iterator=True
            )
        except Exception as error:
            tqdm.write(f"[ERROR] Не удалось открыть файл чанками: {error}")
            return False

        xlsx_path = csv_path.with_suffix(".xlsx")
        sheet_index = 1
        current_rows_in_sheet = 0
        current_sheet_name = f"Sheet{sheet_index}"

        try:
            with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as excel_writer:
                # Прогресс-бар по строкам
                with tqdm(total=estimated_rows, desc=csv_path.name, unit="rows", ncols=120) as progress_bar:
                    for chunk in chunk_iterator:
                        # Создаем новый лист при превышении лимита Excel
                        if current_rows_in_sheet + len(chunk) > EXCEL_MAX_ROWS:
                            sheet_index += 1
                            current_sheet_name = f"Sheet{sheet_index}"
                            current_rows_in_sheet = 0
                            progress_bar.write(f"[INFO] Создаётся новый лист: {current_sheet_name}")

                        # Записываем чанк на текущий лист
                        header = (current_rows_in_sheet == 0)
                        chunk.to_excel(
                            excel_writer,
                            sheet_name=current_sheet_name,
                            index=False,
                            header=header,
                            startrow=current_rows_in_sheet
                        )
                        current_rows_in_sheet += len(chunk)

                        # Обновляем прогресс
                        progress_bar.update(len(chunk))
                        if progress_bar.n % 10000 < len(chunk):
                            progress_bar.set_postfix_str(
                                f"обработано: {progress_bar.n:,} / {estimated_rows:,}"
                            )

                        # Применяем автофильтр после первого чанка на листе
                        if current_rows_in_sheet == len(chunk):
                            worksheet = excel_writer.sheets[current_sheet_name]
                            last_column_index = len(chunk.columns)
                            worksheet.autofilter(0, 0, 0, last_column_index - 1)

                progress_bar.write(f"[INFO] Финальная запись XLSX…")

            tqdm.write(f"[INFO] Сохранён: {xlsx_path.name}")
            return True

        except Exception as error:
            tqdm.write(f"[ERROR] Ошибка при записи большого файла: {error}")
            return False

    def process_single_csv_file(self, csv_path: Path) -> None:
        """
        Обрабатывает один CSV файл.

        Args:
            csv_path: Путь к CSV файлу
        """
        tqdm.write(f"\n[INFO] Обрабатывается файл: {csv_path.name}")

        # Определяем параметры файла
        file_encoding = self.detect_encoding(csv_path)
        file_delimiter = self.detect_delimiter(csv_path, file_encoding)
        estimated_row_count = self.estimate_total_rows(csv_path)

        tqdm.write(f"[INFO] Кодировка: {file_encoding}")
        tqdm.write(f"[INFO] Разделитель: {repr(file_delimiter)}")
        tqdm.write(f"[INFO] Оценочное количество строк: {estimated_row_count:,}")

        # Выбираем метод обработки в зависимости от размера файла
        is_small_file = estimated_row_count <= CHUNK_SIZE * 3

        if is_small_file:
            success = self.process_small_csv_file(csv_path, file_encoding, file_delimiter)
        else:
            success = self.process_large_csv_file(csv_path, file_encoding, file_delimiter, estimated_row_count)

        # Обновляем статистику
        if success:
            self.stats["processed_files"] += 1
        else:
            self.stats["failed_files"].append(csv_path.name)

    def scan_and_process_directory(self, directory: Path) -> None:
        """
        Сканирует директорию и обрабатывает все CSV файлы.

        Args:
            directory: Путь к директории для сканирования
        """
        csv_files = sorted(directory.glob("*.csv"))

        tqdm.write(f"[INFO] Найдено CSV-файлов: {len(csv_files)}")

        if csv_files:
            tqdm.write("=== Список файлов ===")
            for file in csv_files:
                tqdm.write(f" • {file.name}")
            tqdm.write("=====================")

        # Обрабатываем каждый CSV файл
        for csv_file in csv_files:
            self.process_single_csv_file(csv_file)

        self._print_final_statistics()

    def _print_final_statistics(self) -> None:
        """Выводит итоговую статистику обработки."""
        tqdm.write("\n=== Итог анализа ===")
        tqdm.write(f"Успешно обработано: {self.stats['processed_files']}")

        if self.stats["failed_files"]:
            tqdm.write("Ошибки в файлах:")
            for failed_file in self.stats["failed_files"]:
                tqdm.write(f"  - {failed_file}")
        else:
            tqdm.write("Ошибок нет.")


def main() -> None:
    """Основная функция программы."""
    current_directory = Path(".")
    processor = CSVProcessor()
    processor.scan_and_process_directory(current_directory)


if __name__ == "__main__":
    main()