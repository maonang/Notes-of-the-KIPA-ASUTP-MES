#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import sys
import logging
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill
from tqdm import tqdm
import difflib
import pandas as pd
import warnings
import shutil

# Подавление предупреждения openpyxl о Data Validation extension
warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

# ========== НАСТРОЙКИ ==========
TEXT_HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
EXCLUDE_OUTPUTS = {"primary_text_diff.xlsx", "secondary_text_diff.xlsx", "result_analyze.log", "summary_diffs.xlsx"}
FORMAT_STRICTNESS = 'strict'  # 'lenient' or 'strict'
# ==============================

# Основной лог — теперь минимальный, подробные логи идут в отдельные файлы per-sheet
logging.basicConfig(level=logging.INFO, format="%(message)s")


def find_excel_files_in_cwd():
    p = Path('.')
    files = [f for f in sorted(p.glob("*.*")) if f.suffix.lower() in ('.xlsx', '.xlsm')]
    files = [f for f in files if f.name not in EXCLUDE_OUTPUTS and not f.name.startswith('~$')]
    return files

def is_formula_cell(cell):
    try:
        if cell is None:
            return False
        if getattr(cell, "data_type", None) == 'f':
            return True
        v = cell.value
        return isinstance(v, str) and v.startswith('=')
    except Exception:
        return False

def stringify_cell_value(cell):
    if cell is None:
        return ""
    v = cell.value
    if v is None:
        return ""
    if is_formula_cell(cell):
        return str(v).strip()
    return str(v).strip()

def row_signature(sheet, row_idx, min_col=None, max_col=None):
    if min_col is None or max_col is None:
        row = list(sheet[row_idx])
        if not row:
            return ""
        min_c = 1
        max_c = max([c.column for c in row])
    else:
        min_c = min_col; max_c = max_col
    parts = []
    for c in range(min_c, max_c+1):
        parts.append(stringify_cell_value(sheet.cell(row=row_idx, column=c)))
    return "\\x1f".join(parts)

def sequence_similarity(a, b):
    if a is None: a = ""
    if b is None: b = ""
    return difflib.SequenceMatcher(None, a, b).ratio()

def needleman_wunsch_align(seqA, seqB, match_score_func, gap_penalty=-0.2):
    n = len(seqA); m = len(seqB)
    score = [[0.0]*(m+1) for _ in range(n+1)]
    pointer = [[None]*(m+1) for _ in range(n+1)]
    for i in range(1, n+1):
        score[i][0] = score[i-1][0] + gap_penalty; pointer[i][0] = 1
    for j in range(1, m+1):
        score[0][j] = score[0][j-1] + gap_penalty; pointer[0][j] = 2
    for i in range(1, n+1):
        a = seqA[i-1]
        for j in range(1, m+1):
            b = seqB[j-1]
            match = score[i-1][j-1] + match_score_func(a, b)
            delete = score[i-1][j] + gap_penalty
            insert = score[i][j-1] + gap_penalty
            best = match; ptr = 0
            if delete > best:
                best = delete; ptr = 1
            if insert > best:
                best = insert; ptr = 2
            score[i][j] = best; pointer[i][j] = ptr
    i, j = n, m
    alignment = []
    while i > 0 or j > 0:
        ptr = pointer[i][j]
        if ptr == 0:
            alignment.append((seqA[i-1], seqB[j-1])); i -= 1; j -= 1
        elif ptr == 1:
            alignment.append((seqA[i-1], None)); i -= 1
        elif ptr == 2:
            alignment.append((None, seqB[j-1])); j -= 1
        else:
            if i > 0:
                alignment.append((seqA[i-1], None)); i -= 1
            elif j > 0:
                alignment.append((None, seqB[j-1])); j -= 1
    alignment.reverse()
    return alignment

def get_used_bounds(sheet):
    min_row = None; min_col = None; max_row = 0; max_col = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                r, c = cell.row, cell.column
                if min_row is None or r < min_row: min_row = r
                if min_col is None or c < min_col: min_col = c
                if r > max_row: max_row = r
                if c > max_col: max_col = c
    if min_row is None:
        return 1,1,0,0
    return min_row, min_col, max_row, max_col

def format_signature(cell):
    if cell is None: return None
    try:
        font = {
            'name': getattr(cell.font, 'name', None),
            'size': getattr(cell.font, 'sz', None),
            'bold': getattr(cell.font, 'bold', None),
            'italic': getattr(cell.font, 'italic', None),
            'underline': getattr(cell.font, 'underline', None),
            'color': getattr(cell.font.color, None).rgb if getattr(cell.font, 'color', None) and getattr(cell.font.color, 'rgb', None) else None
        }
    except Exception:
        font = None
    try:
        fill = {
            'patternType': getattr(cell.fill, 'patternType', None),
            'fgColor': getattr(cell.fill, 'fgColor', None).rgb if getattr(cell.fill, 'fgColor', None) and getattr(cell.fill.fgColor, 'rgb', None) else None,
            'bgColor': getattr(cell.fill, 'bgColor', None).rgb if getattr(cell.fill, 'bgColor', None) and getattr(cell.fill.bgColor, 'rgb', None) else None
        }
    except Exception:
        fill = None
    try:
        border = {}
        for side_name in ('left','right','top','bottom'):
            side = getattr(cell.border, side_name, None)
            if side is not None:
                border[side_name] = {
                    'style': getattr(side, 'style', None),
                    'color': getattr(side.color, 'rgb', None) if getattr(side, 'color', None) else None
                }
            else:
                border[side_name] = None
    except Exception:
        border = None
    try:
        alignment = {
            'horizontal': getattr(cell.alignment, 'horizontal', None),
            'vertical': getattr(cell.alignment, 'vertical', None),
            'wrapText': getattr(cell.alignment, 'wrap_text', None)
        }
    except Exception:
        alignment = None
    try:
        nf = getattr(cell, 'number_format', None)
    except Exception:
        nf = None
    return {'font': font, 'fill': fill, 'border': border, 'alignment': alignment, 'number_format': nf}

def compare_formats(sig_a, sig_b, strictness='strict'):
    if sig_a is None and sig_b is None: return True, ""
    if sig_a is None or sig_b is None: return False, "одно оформление отсутствует"
    diffs = []
    fa = sig_a.get('font'); fb = sig_b.get('font')
    if fa or fb:
        for k in ('name','size','bold','italic','underline'):
            if (fa.get(k) if fa else None) != (fb.get(k) if fb else None):
                diffs.append(f"font.{k}: {fa.get(k) if fa else None} != {fb.get(k) if fb else None}")
        ca = fa.get('color') if fa else None; cb = fb.get('color') if fb else None
        if ca != cb:
            diffs.append(f"font.color: {ca} != {cb}")
    fla = sig_a.get('fill'); flb = sig_b.get('fill')
    if fla or flb:
        if (fla.get('patternType') if fla else None) != (flb.get('patternType') if flb else None):
            diffs.append(f"fill.patternType: {fla.get('patternType') if fla else None} != {flb.get('patternType') if flb else None}")
        fa_fg = fla.get('fgColor') if fla else None; fb_fg = flb.get('fgColor') if flb else None
        if fa_fg != fb_fg:
            diffs.append(f"fill.fgColor: {fa_fg} != {fb_fg}")
    ba = sig_a.get('border'); bb = sig_b.get('border')
    if ba or bb:
        for side in ('left','right','top','bottom'):
            sa = ba.get(side) if ba else None; sb = bb.get(side) if bb else None
            if sa != sb:
                diffs.append(f"border.{side}: {sa} != {sb}")
    aa = sig_a.get('alignment'); ab = sig_b.get('alignment')
    if aa != ab:
        diffs.append(f"alignment: {aa} != {ab}")
    if (sig_a.get('number_format') != sig_b.get('number_format')):
        diffs.append(f"number_format: {sig_a.get('number_format')} != {sig_b.get('number_format')}")
    equal = len(diffs) == 0
    return equal, "; ".join(diffs)

def align_and_compare_sheets(ws_primary, ws_secondary):
    diffs = {'Text': [], 'Formulas': [], 'Formatting': [], 'Hidden': {'rows_added': [], 'rows_removed': [], 'cols_added': [], 'cols_removed': []}, 'Comments': []}
    pmin_r, pmin_c, pmax_r, pmax_c = get_used_bounds(ws_primary)
    smin_r, smin_c, smax_r, smax_c = get_used_bounds(ws_secondary)
    if pmax_r == 0 and smax_r == 0:
        return diffs, []
    prim_rows = list(range(pmin_r, pmax_r+1))
    sec_rows = list(range(smin_r, smax_r+1))
    prim_sigs = [row_signature(ws_primary, r, pmin_c, pmax_c) for r in prim_rows]
    sec_sigs = [row_signature(ws_secondary, r, smin_c, smax_c) for r in sec_rows]
    def row_score(i,j):
        return sequence_similarity(prim_sigs[i], sec_sigs[j])
    seqA = prim_rows; seqB = sec_rows
    idx_map_A = {r:i for i,r in enumerate(seqA)}; idx_map_B = {r:i for i,r in enumerate(seqB)}
    def match_wrapper_fast(a,b):
        i = idx_map_A[a]; j = idx_map_B[b]; return row_score(i,j)
    row_alignment = needleman_wunsch_align(seqA, seqB, match_wrapper_fast, gap_penalty=-0.15)
    aligned_rows = row_alignment
    mapping_positions = []
    for pair in tqdm(aligned_rows, desc="    выравнивание строк", leave=False):
        p_row, s_row = pair
        if p_row is None:
            diffs['Hidden']['rows_added'].append(s_row); mapping_positions.append((None, s_row, [])); continue
        if s_row is None:
            diffs['Hidden']['rows_removed'].append(p_row); mapping_positions.append((p_row, None, [])); continue
        prim_cols = list(range(pmin_c, pmax_c+1)); sec_cols = list(range(smin_c, smax_c+1))
        prim_cells = [stringify_cell_value(ws_primary.cell(row=p_row, column=c)) for c in prim_cols]
        sec_cells = [stringify_cell_value(ws_secondary.cell(row=s_row, column=c)) for c in sec_cols]
        prim_idx_map = {c:i for i,c in enumerate(prim_cols)}; sec_idx_map = {c:i for i,c in enumerate(sec_cols)}
        def match_col_fast(a,b):
            i = prim_idx_map[a]; j = sec_idx_map[b]; return sequence_similarity(prim_cells[i], sec_cells[j])
        col_alignment = needleman_wunsch_align(prim_cols, sec_cols, match_col_fast, gap_penalty=-0.3)
        aligned_cols = col_alignment
        mapping_positions.append((p_row, s_row, aligned_cols))
        for p_col, s_col in aligned_cols:
            p_cell = ws_primary.cell(row=p_row, column=p_col) if p_col is not None else None
            s_cell = ws_secondary.cell(row=s_row, column=s_col) if s_col is not None else None
            p_coord = f"{get_column_letter(p_col)}{p_row}" if p_col is not None else None
            s_coord = f"{get_column_letter(s_col)}{s_row}" if s_col is not None else None
            p_is_formula = is_formula_cell(p_cell); s_is_formula = is_formula_cell(s_cell)
            if p_is_formula or s_is_formula:
                p_f = str(p_cell.value).strip() if p_cell and p_cell.value is not None else ""
                s_f = str(s_cell.value).strip() if s_cell and s_cell.value is not None else ""
                if p_f != s_f:
                    diffs['Formulas'].append((p_coord, s_coord, p_f, s_f))
                continue
            p_text = "" if p_cell is None or p_cell.value is None else str(p_cell.value).strip()
            s_text = "" if s_cell is None or s_cell.value is None else str(s_cell.value).strip()
            if p_text != s_text:
                if not (p_text == "" and s_text == ""):
                    diffs['Text'].append((p_coord, s_coord, p_text, s_text))
            sig_p = format_signature(p_cell); sig_s = format_signature(s_cell)
            equal_fmt, details = compare_formats(sig_p, sig_s, strictness=FORMAT_STRICTNESS)
            if not equal_fmt:
                diffs['Formatting'].append((p_coord, s_coord, details))
            p_comm = p_cell.comment.text if (p_cell is not None and getattr(p_cell, 'comment', None)) else None
            s_comm = s_cell.comment.text if (s_cell is not None and getattr(s_cell, 'comment', None)) else None
            if (p_comm or s_comm) and p_comm != s_comm:
                diffs['Comments'].append((p_coord, s_coord, p_comm, s_comm))
    return diffs, mapping_positions

# --- Логгеры per sheet ---
def prepare_loggers(sheet_name, root_dir):
    sheet_dir = Path(root_dir) / sheet_name
    sheet_dir.mkdir(parents=True, exist_ok=True)
    files = {
        'text': open(sheet_dir / "result - Текст.log", 'w', encoding='utf-8'),
        'formula': open(sheet_dir / "result - Формулы.log", 'w', encoding='utf-8'),
        'style': open(sheet_dir / "result - Оформление.log", 'w', encoding='utf-8'),
        'hidden': open(sheet_dir / "result - Сокрытие строк, столбцов.log", 'w', encoding='utf-8'),
        'comments': open(sheet_dir / "result - Добавление, удаление примечаний.log", 'w', encoding='utf-8'),
        'dir': sheet_dir
    }
    return files

def close_loggers(files):
    for k, f in files.items():
        if hasattr(f, 'close') and k != 'dir':
            try:
                f.close()
            except Exception:
                pass

def save_summary_excel(diffs, out_path):
    writer = pd.ExcelWriter(out_path, engine='openpyxl')
    any_data = False
    # Text
    text_rows = []
    for p_coord, s_coord, p_text, s_text in diffs.get('Text', []):
        text_rows.append({'p_coord': p_coord, 's_coord': s_coord, 'p_text': p_text, 's_text': s_text})
    if text_rows:
        any_data = True
        pd.DataFrame(text_rows).to_excel(writer, sheet_name='Text', index=False)
    # Formulas
    form_rows = []
    for p_coord, s_coord, p_f, s_f in diffs.get('Formulas', []):
        form_rows.append({'p_coord': p_coord, 's_coord': s_coord, 'p_formula': p_f, 's_formula': s_f})
    if form_rows:
        any_data = True
        pd.DataFrame(form_rows).to_excel(writer, sheet_name='Formulas', index=False)
    # Formatting
    fmt_rows = []
    for p_coord, s_coord, details in diffs.get('Formatting', []):
        fmt_rows.append({'p_coord': p_coord, 's_coord': s_coord, 'details': details})
    if fmt_rows:
        any_data = True
        pd.DataFrame(fmt_rows).to_excel(writer, sheet_name='Formatting', index=False)
    # Hidden
    hidden = diffs.get('Hidden', {})
    if hidden and (hidden.get('rows_added') or hidden.get('rows_removed') or hidden.get('cols_added') or hidden.get('cols_removed')):
        any_data = True
        pd.DataFrame([hidden]).to_excel(writer, sheet_name='Hidden', index=False)
    # Comments
    comm_rows = []
    for p_coord, s_coord, p_c, s_c in diffs.get('Comments', []):
        comm_rows.append({'p_coord': p_coord, 's_coord': s_coord, 'p_comment': p_c, 's_comment': s_c})
    if comm_rows:
        any_data = True
        pd.DataFrame(comm_rows).to_excel(writer, sheet_name='Comments', index=False)
    if not any_data:
        pd.DataFrame([{'info':'No differences detected'}]).to_excel(writer, sheet_name='No_diffs', index=False)
    # Important: close writer (writer.save() may not exist in some pandas versions)
    writer.close()

def create_single_sheet_copy_with_highlights(src_path, sheet_name, highlights_set, out_path):
    keep_vba = src_path.suffix.lower() == '.xlsm'
    wb = load_workbook(src_path, data_only=False, keep_vba=keep_vba)
    # If sheet not present, create empty workbook with that name
    if sheet_name not in wb.sheetnames:
        # create a workbook with single empty sheet
        new_wb = Workbook()
        ws = new_wb.active
        ws.title = sheet_name
        new_wb.save(out_path)
        return
    # Remove other sheets
    for name in list(wb.sheetnames):
        if name != sheet_name:
            try:
                wb.remove(wb[name])
            except Exception:
                pass
    ws = wb[sheet_name]
    # Apply highlight to provided coordinates
    for coord in highlights_set:
        try:
            cell = ws[coord]
            # preserve existing fill (this is the visual result of conditional formatting if already materialized)
            cell.fill = TEXT_HIGHLIGHT_FILL
        except Exception:
            pass
    # Remove conditional formatting rules but keep cell.fill values (already present)
    try:
        # Some versions have clear method; otherwise set internal dict to {}
        if hasattr(ws.conditional_formatting, 'clear'):
            ws.conditional_formatting.clear()
        else:
            ws.conditional_formatting._cf_rules = {}
    except Exception:
        try:
            ws.conditional_formatting._cf_rules = {}
        except Exception:
            pass
    # Save keeping vba if needed
    wb.save(out_path)

def write_logs_to_files(diffs, loggers):
    # Write Text
    f = loggers['text']
    if diffs.get('Text'):
        f.write("Категория: Текст (p_coord <-> s_coord)\n")
        for p_coord, s_coord, p_text, s_text in diffs['Text']:
            f.write(f"{p_coord} <-> {s_coord}\n- {p_text}\n- {s_text}\n\n")
    else:
        f.write("Категория: Текст — различий не обнаружено.\n")
    # Formulas
    f = loggers['formula']
    if diffs.get('Formulas'):
        f.write("Категория: Формулы (p_coord <-> s_coord)\n")
        for p_coord, s_coord, p_f, s_f in diffs['Formulas']:
            f.write(f"{p_coord} <-> {s_coord}\n- {p_f}\n- {s_f}\n\n")
    else:
        f.write("Категория: Формулы — различий не обнаружено.\n")
    # Formatting
    f = loggers['style']
    if diffs.get('Formatting'):
        f.write("Категория: Оформление (p_coord <-> s_coord)\n")
        for p_coord, s_coord, details in diffs['Formatting']:
            f.write(f"{p_coord} <-> {s_coord}\n- {details}\n\n")
    else:
        f.write("Категория: Оформление — различий не обнаружено.\n")
    # Hidden
    f = loggers['hidden']
    hidden = diffs.get('Hidden', {})
    f.write("Категория: Сокрытие строк/столбцов\n")
    f.write(f"  строки добавлены во вторичном: {hidden.get('rows_added')}\n")
    f.write(f"  строки удалены во вторичном: {hidden.get('rows_removed')}\n")
    f.write(f"  cols_added: {hidden.get('cols_added')}; cols_removed: {hidden.get('cols_removed')}\n")
    # Comments
    f = loggers['comments']
    if diffs.get('Comments'):
        f.write("Категория: Примечания (p_coord <-> s_coord)\n")
        for p_coord, s_coord, p_c, s_c in diffs['Comments']:
            f.write(f"{p_coord} <-> {s_coord}\n- {p_c}\n- {s_c}\n\n")
    else:
        f.write("Категория: Примечания — различий не обнаружено.\n")

def process_sheet_pair(primary_path, secondary_path, sheet_name, out_root):
    loggers = prepare_loggers(sheet_name, out_root)
    try:
        wb_p = load_workbook(primary_path, data_only=False, keep_vba=(primary_path.suffix.lower()=='.xlsm'))
        wb_s = load_workbook(secondary_path, data_only=False, keep_vba=(secondary_path.suffix.lower()=='.xlsm'))
    except Exception as e:
        logging.info(f"Ошибка при загрузке файлов: {e}")
        close_loggers(loggers)
        return
    ws_p = wb_p[sheet_name] if sheet_name in wb_p.sheetnames else None
    ws_s = wb_s[sheet_name] if sheet_name in wb_s.sheetnames else None
    if ws_p is None and ws_s is None:
        logging.info(f"Лист {sheet_name} отсутствует в обоих файлах — пропуск.")
        close_loggers(loggers); return
    if ws_p is None:
        logging.info(f"Лист {sheet_name} отсутствует в основном файле — помечу как добавленный во вторичном.")
    if ws_s is None:
        logging.info(f"Лист {sheet_name} отсутствует во вторичном файле — помечу как удалён.")
    diffs, mapping = align_and_compare_sheets(ws_p or Workbook().active, ws_s or Workbook().active)
    # Запись логов
    write_logs_to_files(diffs, loggers)
    # Собираем координаты для подсветки текстовых отличий
    prim_coords = set()
    sec_coords = set()
    for p_coord, s_coord, p_text, s_text in diffs['Text']:
        if p_coord: prim_coords.add(p_coord)
        if s_coord: sec_coords.add(s_coord)
    # Создаём копии с одним листом в папке
    sheet_dir = loggers['dir']
    out_p = sheet_dir / (f"primary_text_diff{'.xlsm' if primary_path.suffix.lower()=='.xlsm' else '.xlsx'}")
    out_s = sheet_dir / (f"secondary_text_diff{'.xlsm' if secondary_path.suffix.lower()=='.xlsm' else '.xlsx'}")
    create_single_sheet_copy_with_highlights(primary_path, sheet_name, prim_coords, out_p)
    create_single_sheet_copy_with_highlights(secondary_path, sheet_name, sec_coords, out_s)
    # Сохраняем summary_diffs.xlsx в папке листа
    summary_path = sheet_dir / 'summary_diffs.xlsx'
    try:
        save_summary_excel(diffs, str(summary_path))
        logging.info(f"Сводная таблица для листа '{sheet_name}' сохранена в: {summary_path}")
    except Exception as e:
        logging.info(f"Не удалось сохранить summary для листа {sheet_name}: {e}")
    close_loggers(loggers)

def main():
    files = find_excel_files_in_cwd()
    if len(files) < 2:
        print("В текущем каталоге меньше двух .xlsx/.xlsm файлов. Положите хотя бы два файла и запустите снова.")
        return
    print("Найдены Excel файлы:")
    for i,f in enumerate(files, start=1):
        print(f"  {i}. {f.name}")
    try:
        idx1 = int(input("Выберите номер первого файла (по умолчанию 1): ") or "1")
        idx2 = int(input("Выберите номер второго файла (по умолчанию 2): ") or "2")
        if idx1 < 1 or idx2 < 1 or idx1 > len(files) or idx2 > len(files) or idx1 == idx2:
            print("Неверный выбор — используем первые два по умолчанию.")
            idx1, idx2 = 1, 2
    except Exception:
        idx1, idx2 = 1, 2
    file1 = files[idx1-1]; file2 = files[idx2-1]
    print(f"1) {file1.name}\n2) {file2.name}")
    pick = input("Какой файл считать основным? Введите 1 или 2 (по умолчанию 1): ") or "1"
    if pick.strip() == "2":
        primary_path = file2; secondary_path = file1
    else:
        primary_path = file1; secondary_path = file2
    # Загружаем книги (для определения листов)
    wb_p = load_workbook(primary_path, data_only=False, keep_vba=(primary_path.suffix.lower()=='.xlsm'))
    wb_s = load_workbook(secondary_path, data_only=False, keep_vba=(secondary_path.suffix.lower()=='.xlsm'))
    sheets_p = wb_p.sheetnames; sheets_s = wb_s.sheetnames
    # Если набор имён листов совпадает — анализируем все без запроса
    set_p = set(sheets_p); set_s = set(sheets_s)
    out_root = Path('./results')
    out_root.mkdir(exist_ok=True)
    if set_p == set_s:
        sheets_to_process = sorted(list(set_p))
        logging.info("Имена листов совпадают — анализ всех листов автоматически.")
    else:
        # спрошим пользователя для выбора листа в каждом файле (как раньше)
        print("\nЛисты в основном файле:")
        for i,name in enumerate(sheets_p, start=1):
            print(f"  {i}. {name}")
        try:
            sidx_p = int(input("Выберите номер листа в основном файле для анализа (по умолчанию 1): ") or "1")
            if sidx_p < 1 or sidx_p > len(sheets_p): sidx_p = 1
        except Exception:
            sidx_p = 1
        sheet_primary_name = sheets_p[sidx_p-1]
        print("\nЛисты во вторичном файле:")
        for i,name in enumerate(sheets_s, start=1):
            print(f"  {i}. {name}")
        try:
            sidx_s = int(input("Выберите номер листа во вторичном файле для анализа (по умолчанию 1): ") or "1")
            if sidx_s < 1 or sidx_s > len(sheets_s): sidx_s = 1
        except Exception:
            sidx_s = 1
        sheet_secondary_name = sheets_s[sidx_s-1]
        # If selected names are same, process single; otherwise process the two names (intersection handled later)
        sheets_to_process = [sheet_primary_name, sheet_secondary_name]
    # Для каждого листа запускаем процессинг
    for sheet_name in sheets_to_process:
        process_sheet_pair(primary_path, secondary_path, sheet_name, out_root)
    print("Готово. Смотрите папку ./results для логов и diff-файлов по листам.")

if __name__ == '__main__':
    main()